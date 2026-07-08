"""stocks.json 저장 위치, 자동 마이그레이션, Excel import/export."""

import os
import re
import sys
import uuid
import shutil
from pathlib import Path

from .portfolio import portfolio_totals, stock_metrics


# ─── 설정 파일 경로 (OS별 표준 디렉토리) ──────────────────────────────────────
def _config_dir() -> Path:
    if sys.platform == "win32":
        base = Path(os.environ.get("APPDATA") or (Path.home() / "AppData" / "Roaming"))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path(os.environ.get("XDG_CONFIG_HOME") or (Path.home() / ".config"))
    d = base / "Pinstock"
    d.mkdir(parents=True, exist_ok=True)
    return d


CONFIG_FILE = str(_config_dir() / "stocks.json")
BACKUP_FILE = CONFIG_FILE + ".bak"


# ─── 종목 스키마 기본값 ─────────────────────────────────────────────────────
MARKET_KR = "KR"
MARKET_US = "US"
CURRENCY_KRW = "KRW"
CURRENCY_USD = "USD"

# ─── 관심종목 태그 기본값 ─────────────────────────────────────────────────────
# 태그는 관심종목 전용이며 종목당 1개만 부여한다. 색상은 필수(#rrggbb).
DEFAULT_TAG_COLOR = "#89b4fa"
_HEX_COLOR_RE = re.compile(r"^#[0-9a-fA-F]{6}$")


def normalize_stock_schema(stock: dict) -> dict:
    """기존 stocks.json 항목에 시장/통화 기본값을 보강한다.

    미국 주식 지원 전의 기존 데이터는 market/currency 필드가 없으므로 한국
    주식으로 취급한다. 알 수 없는 부가 필드(pos, hidden 등)는 그대로 보존한다.
    """
    normalized = dict(stock)
    market = str(normalized.get("market") or MARKET_KR).strip().upper()
    if market not in {MARKET_KR, MARKET_US}:
        market = MARKET_KR
    normalized["market"] = market

    default_currency = CURRENCY_USD if market == MARKET_US else CURRENCY_KRW
    currency = str(normalized.get("currency") or default_currency).strip().upper()
    normalized["currency"] = currency or default_currency

    if market == MARKET_US and "buy_exchange_rate" in normalized:
        try:
            normalized["buy_exchange_rate"] = float(normalized["buy_exchange_rate"])
        except (TypeError, ValueError):
            normalized.pop("buy_exchange_rate", None)

    return normalized


def normalize_stocks_schema(stocks: list[dict]) -> list[dict]:
    return [normalize_stock_schema(s) for s in stocks if isinstance(s, dict)]


# ─── 메모(투자 메모장) 스키마 ─────────────────────────────────────────────────
# 앱 전체에서 1개만 존재하는 자유 메모. {text, updated_at} 객체로 저장하며,
# 구버전/형식 깨짐(문자열·None 등)도 안전하게 받아 준다.
def normalize_memo(raw) -> dict:
    """메모를 {text: str, updated_at: str|None, geometry: [x,y,w,h]|None} 로 정규화.

    geometry 는 메모창의 마지막 위치/크기. 구버전(문자열 형태·geometry 없음)도
    안전하게 받아 준다.
    """
    text, updated, geometry = "", None, None
    if isinstance(raw, str):
        text = raw
    elif isinstance(raw, dict):
        t = raw.get("text")
        text = t if isinstance(t, str) else ""
        u = raw.get("updated_at")
        updated = u if isinstance(u, str) else None
        g = raw.get("geometry")
        if isinstance(g, (list, tuple)) and len(g) == 4:
            try:
                geometry = [int(g[0]), int(g[1]), int(g[2]), int(g[3])]
            except (TypeError, ValueError):
                geometry = None
    return {"text": text, "updated_at": updated, "geometry": geometry}


# ─── 종목별 메모 저장소 스키마 ────────────────────────────────────────────────
# 종목 코드로 keyed 된 별도 저장소 {code: {text, updated_at, geometry, name}}.
# 메모를 종목 dict 안이 아니라 여기 따로 두므로, 종목을 삭제해도 메모는 남고
# 같은 코드의 종목을 다시 추가하면 코드 매칭으로 자동으로 다시 연결된다.
# name 은 종목이 삭제된 뒤에도 목록에 이름을 보여주기 위한 마지막 종목명이다.
def normalize_stock_memos(raw) -> dict:
    """종목별 메모 저장소를 {code: {text, updated_at, geometry, name}} 로 정규화.

    텍스트가 빈 항목은 '메모 없음'으로 보고 버린다(메모창에서 텍스트를 다 지우면
    삭제되는 동작과 일치). code/형식이 이상한 항목도 안전하게 걸러 낸다.
    """
    out: dict[str, dict] = {}
    if not isinstance(raw, dict):
        return out
    for code, m in raw.items():
        if not isinstance(code, str) or not code:
            continue
        base = normalize_memo(m)   # {text, updated_at, geometry}
        if not (base.get("text") or "").strip():
            continue
        name = m.get("name") if isinstance(m, dict) else None
        base["name"] = name.strip() if isinstance(name, str) else ""
        out[code] = base
    return out


# ─── 분리(detach) 창 상태 스키마 (macOS 전용) ─────────────────────────────────
# 보유/관심 중 한 뷰를 독립 창으로 분리한 상태. 어느 뷰가 분리됐는지(view)와 그
# 창의 위치/높이/고정/투명도/시장필터를 저장해 재시작 후 복원한다. 형식이 깨졌거나
# 구버전(없음)이면 분리 안 한 기본 상태로 돌린다.
def normalize_detached(raw) -> dict:
    """{view: 'holdings'|'watch'|None, pos: [x,y]|None, height: int|None,
    pinned: bool, opacity: float, market_filter: 'ALL'|'KR'|'US'} 로 정규화."""
    view = None
    pos = None
    height = None
    pinned = False
    opacity = 1.0
    market_filter = "ALL"
    if isinstance(raw, dict):
        v = raw.get("view")
        if v in ("holdings", "watch"):
            view = v
        p = raw.get("pos")
        if isinstance(p, (list, tuple)) and len(p) == 2:
            try:
                pos = [int(p[0]), int(p[1])]
            except (TypeError, ValueError):
                pos = None
        h = raw.get("height")
        if h is not None:
            try:
                height = int(h)
            except (TypeError, ValueError):
                height = None
        pinned = bool(raw.get("pinned", True))
        try:
            opacity = max(0.1, min(1.0, float(raw.get("opacity", 1.0))))
        except (TypeError, ValueError):
            opacity = 1.0
        mf = str(raw.get("market_filter") or "ALL").strip().upper()
        market_filter = mf if mf in ("ALL", "KR", "US") else "ALL"
    return {"view": view, "pos": pos, "height": height,
            "pinned": pinned, "opacity": opacity, "market_filter": market_filter}


# ─── 관심종목(워치리스트) 스키마 ──────────────────────────────────────────────
# 관심종목은 보유와 완전히 독립된 별도 목록이다. 평단가/수량/손익 개념이 없고,
# 시세는 일봉 기준으로 본다. 같은 종목이 보유와 관심에 동시에 존재할 수 있다.
def normalize_watch_item(item: dict) -> dict:
    """관심종목 한 항목에 시장/통화 기본값을 보강하고 관심 전용 필드를 정규화한다.

    보유 종목과 달리 avg_price/quantity 는 없다. tag(태그 레지스트리 id, 종목당
    1개), hidden(표시 ON/OFF), pos(위젯 위치)는 있으면 보존한다.
    """
    normalized = dict(item)

    market = str(normalized.get("market") or MARKET_KR).strip().upper()
    if market not in {MARKET_KR, MARKET_US}:
        market = MARKET_KR
    normalized["market"] = market

    default_currency = CURRENCY_USD if market == MARKET_US else CURRENCY_KRW
    currency = str(normalized.get("currency") or default_currency).strip().upper()
    normalized["currency"] = currency or default_currency

    # 타입: 'index'(지수) 또는 'stock'(개별 종목). 지수는 시세 라우팅·표시 포맷이
    # 다르므로 구분해 보존한다. 알 수 없는 값은 종목으로 본다.
    item_type = str(normalized.get("type") or "").strip().lower()
    normalized["type"] = "index" if item_type == "index" else "stock"

    # 태그: 종목당 1개. 태그 레지스트리(watch_tags)의 id 를 참조한다. 없으면 "".
    tag = normalized.get("tag")
    normalized["tag"] = tag.strip() if isinstance(tag, str) else ""
    normalized.pop("tags", None)   # 구버전(리스트형) 필드 정리

    normalized["hidden"] = bool(normalized.get("hidden", False))
    return normalized


def normalize_watchlist_schema(items: list[dict]) -> list[dict]:
    return [normalize_watch_item(i) for i in items if isinstance(i, dict)]


# ─── 관심종목 태그 레지스트리 ─────────────────────────────────────────────────
# 태그는 {id, name, color} 로 저장하고, 관심종목 항목은 tag(=id)로 참조한다.
# id 는 이름/색상을 바꿔도 참조가 끊기지 않도록 한 번 만들면 고정한다.
def new_tag_id() -> str:
    return uuid.uuid4().hex[:8]


def normalize_tag(tag: dict) -> dict | None:
    """태그 한 개를 정규화. id/name 이 비면 None(무효). 색상이 이상하면 기본색."""
    if not isinstance(tag, dict):
        return None
    tid = str(tag.get("id") or "").strip()
    name = str(tag.get("name") or "").strip()
    if not tid or not name:
        return None
    color = str(tag.get("color") or "").strip()
    if not _HEX_COLOR_RE.match(color):
        color = DEFAULT_TAG_COLOR
    return {"id": tid, "name": name, "color": color.lower()}


def normalize_tags(tags: list) -> list[dict]:
    """태그 레지스트리 정규화. 무효 항목·중복 id 는 제거한다."""
    out: list[dict] = []
    seen: set[str] = set()
    for t in tags or []:
        nt = normalize_tag(t)
        if nt and nt["id"] not in seen:
            seen.add(nt["id"])
            out.append(nt)
    return out


def tag_color_map(tags: list[dict]) -> dict[str, str]:
    """{tag_id: color} 빠른 조회용 매핑."""
    return {t["id"]: t["color"] for t in tags if isinstance(t, dict) and t.get("id")}


def prune_watch_tags(watchlist: list[dict], tags: list[dict]) -> None:
    """레지스트리에 없는 태그 id 를 참조하는 관심종목의 tag 를 비운다(제자리 수정)."""
    valid = {t["id"] for t in tags if isinstance(t, dict) and t.get("id")}
    for w in watchlist:
        if w.get("tag") and w["tag"] not in valid:
            w["tag"] = ""


# ─── 레거시 위치(레포 루트/CWD)에서 새 위치로 1회 자동 이전 ───────────────
def migrate_legacy_config() -> None:
    """저장소 루트(또는 현재 작업 디렉토리)에 있던 stocks.json 을 새 위치로 1회 이전.

    v1.x 시절 단일 스크립트 옆에 stocks.json 을 저장하던 기존 사용자를 위함.
    새 위치에 이미 파일이 있으면 아무것도 하지 않는다.
    이전 후 옛 파일은 같은 경로에 `.migrated` 마커를 남겨 확인할 수 있게 한다.
    """
    if os.path.exists(CONFIG_FILE):
        return
    candidates = [
        # 레포 루트 (pinstock/core/storage.py 의 두 단계 부모)
        Path(__file__).resolve().parent.parent.parent / "stocks.json",
        # 현재 작업 디렉토리
        Path.cwd() / "stocks.json",
    ]
    seen: set[Path] = set()
    for legacy in candidates:
        try:
            legacy = legacy.resolve()
        except Exception:
            continue
        if legacy in seen:
            continue
        seen.add(legacy)
        if legacy.is_file():
            try:
                shutil.move(str(legacy), CONFIG_FILE)
                marker = Path(str(legacy) + ".migrated")
                marker.write_text(
                    f"이 파일은 새 위치로 이동되었습니다:\n{CONFIG_FILE}\n",
                    encoding="utf-8",
                )
                print(f"[migrate] stocks.json 을 {CONFIG_FILE} 로 이전했습니다.")
            except Exception as e:
                print(f"[migrate] 오류: {e}")
            return


# ─── Excel import/export 컬럼 정의 ────────────────────────────────────────────
# 헤더 ↔ stocks.json 필드 매핑. 순서는 export 시 컬럼 순서가 됨.
EXCEL_COLUMNS = [
    ("종목코드", "code"),
    ("종목명",   "name"),
    ("평단가",   "avg_price"),
    ("수량",     "quantity"),
]

EXCEL_OPTIONAL_COLUMNS = [
    ("시장",     "market"),
    ("통화",     "currency"),
    ("매수환율", "buy_exchange_rate"),
]

_US_TICKER_RE = re.compile(r"^\^?[A-Z][A-Z0-9.-]{0,14}$")


def _display_market(market: str) -> str:
    return MARKET_US if str(market or "").strip().upper() == MARKET_US else MARKET_KR


def _normalize_excel_market(value, code: str) -> str:
    raw = str(value or "").strip().upper()
    if raw in {MARKET_US, "USA", "US STOCK", "U.S.", "미국", "미장"}:
        return MARKET_US
    if raw in {MARKET_KR, "KOR", "KOREA", "한국", "국내", "국장"}:
        return MARKET_KR
    if len(code) == 6 and code.isalnum() and any(ch.isdigit() for ch in code):
        return MARKET_KR
    if _US_TICKER_RE.match(code):
        return MARKET_US
    return MARKET_KR


def _normalize_excel_currency(value, market: str) -> str:
    default = CURRENCY_USD if market == MARKET_US else CURRENCY_KRW
    currency = str(value or default).strip().upper()
    return currency or default


# ─── Excel import/export ─────────────────────────────────────────────────────
def export_stocks_to_excel(stocks: list[dict], path: str,
                           current_prices: dict | None = None,
                           usd_krw_rate: float | None = None) -> None:
    """보유 종목을 .xlsx 로 내보내기.
    - 종목코드는 텍스트 셀로 저장 (선행 0/미국 티커 보존: '005930', 'NVDA').
    - 시장/통화/매수환율은 선택 컬럼으로 함께 저장해 미국 주식 라운드트립을 보존한다.
    - 위젯 위치(pos)는 제외 — 다른 PC에서는 화면 좌표가 달라 의미가 없음.
    - current_prices ({code: price}) 와 usd_krw_rate 가 주어지면 시트 하단에
      포트폴리오 요약(총 매입금액 / 평가금액 / 평가손익 / 수익률)을 빈 행 한
      줄로 분리해서 추가. 미국 주식은 원화 기준으로 합산한다.
      import 시에는 빈 행 이후의 행을 모두 무시하므로 라운드트립에 영향 없음."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "보유종목"

    # 종목별 수익 금액/수익률은 현재가·환율로 계산되는 값이라 EXCEL_COLUMNS(=import
    # 필수 컬럼)에는 넣지 않고 export 시에만 뒤에 덧붙인다. import 은 추가 컬럼을
    # 무시하므로 라운드트립에 영향 없음.
    export_columns = EXCEL_COLUMNS + EXCEL_OPTIONAL_COLUMNS
    headers = [h for h, _ in export_columns] + ["수익금액 (원)", "수익률 (%)"]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    prices = current_prices or {}
    for s in stocks:
        row = []
        normalized = normalize_stock_schema(s)
        market = normalized["market"]
        for _, key in export_columns:
            if key == "code":
                row.append(str(normalized.get("code", "")))
            elif key == "name":
                row.append(normalized.get("name", normalized.get("code", "")))
            elif key == "market":
                row.append(_display_market(market))
            elif key == "currency":
                row.append(normalized.get("currency", CURRENCY_USD if market == MARKET_US else CURRENCY_KRW))
            elif key == "buy_exchange_rate":
                row.append(normalized.get("buy_exchange_rate", ""))
            elif key == "quantity":
                row.append(float(normalized.get(key, 0)))
            elif key == "avg_price" and market == MARKET_US:
                row.append(float(normalized.get(key, 0)))
            else:
                row.append(int(normalized.get(key, 0)))
        metrics = stock_metrics(normalized, prices.get(normalized.get("code")), usd_krw_rate)
        row.append(metrics["profit"])
        row.append(round(metrics["profit_rate"], 2))
        ws.append(row)

    # 수익금액/수익률 컬럼 숫자 포맷 (헤더 제외)
    profit_col_idx = len(export_columns) + 1
    rate_col_idx = len(export_columns) + 2
    for row_idx in range(2, ws.max_row + 1):
        pc = ws.cell(row=row_idx, column=profit_col_idx)
        pc.number_format = "#,##0"
        pc.alignment = Alignment(horizontal="right")
        rc = ws.cell(row=row_idx, column=rate_col_idx)
        rc.number_format = "0.00"
        rc.alignment = Alignment(horizontal="right")

    # 종목코드 컬럼을 텍스트 포맷으로 (선행 0/영문 안전)
    code_col_idx = next(i for i, (_, k) in enumerate(export_columns, 1) if k == "code")
    code_letter = ws.cell(row=1, column=code_col_idx).column_letter
    for cell in ws[code_letter][1:]:   # 헤더 제외
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    # 컬럼 너비 자동 조정 (간단히 헤더+여유)
    widths = {"종목코드": 12, "종목명": 28, "평단가": 12, "수량": 10,
              "시장": 8, "통화": 8, "매수환율": 12,
              "수익금액 (원)": 14, "수익률 (%)": 12}
    for col_idx, header in enumerate(headers, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = widths.get(header, 14)

    # ── 포트폴리오 요약 (종목이 1개 이상일 때) ────────────────────────
    # 종목 표와 빈 행 한 줄로 분리. import 측에서 빈 행 이후를 모두 무시하므로
    # 라운드트립 안전.
    if stocks:
        totals = portfolio_totals(
            stocks,
            current_prices=current_prices,
            usd_krw_rate=usd_krw_rate,
            include_hidden=True,
        )
        total_invest = totals["total_invest"]
        total_eval = totals["total_eval"]
        profit = totals["profit"]
        prate = totals["profit_rate"]

        # 빈 행 한 줄 띄우고 다음 행에 요약 헤더
        header_row = ws.max_row + 2
        ws.cell(row=header_row, column=1, value="포트폴리오 요약").font = bold

        rows = [
            ("총 매입금액", total_invest, "#,##0"),
            ("평가금액",   total_eval,   "#,##0"),
            ("평가손익",   profit,        "#,##0"),
            ("수익률 (%)", round(prate, 2), "0.00"),
        ]
        for i, (label, val, fmt) in enumerate(rows, 1):
            r = header_row + i
            ws.cell(row=r, column=1, value=label)
            val_cell = ws.cell(row=r, column=2, value=val)
            val_cell.number_format = fmt
            val_cell.alignment = Alignment(horizontal="right")

    wb.save(path)


def import_stocks_from_excel(path: str) -> list[dict]:
    """Excel 파일에서 보유 종목을 읽어 stocks.json 형식 dict 리스트로 반환.
    검증 실패 시 ValueError 를 발생시킨다 (메시지는 사용자에게 그대로 표시 가능)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        raise ValueError("시트가 비어 있습니다.")

    # 1행 헤더 읽기 (공백/None 안전)
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = [h for h, _ in EXCEL_COLUMNS]
    missing = [h for h in required if h not in header_row]
    if missing:
        raise ValueError(
            "필수 컬럼이 누락되었습니다: " + ", ".join(missing)
            + f"\n(필요한 헤더: {', '.join(required)})"
        )

    # 헤더명 → 컬럼 인덱스. 시장/통화/매수환율은 새 export 에만 있는 선택 컬럼이다.
    optional = [h for h, _ in EXCEL_OPTIONAL_COLUMNS]
    idx_of = {h: header_row.index(h) for h in required if h in header_row}
    idx_of.update({h: header_row.index(h) for h in optional if h in header_row})

    stocks: list[dict] = []
    seen_codes: set[str] = set()
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 빈 행을 만나면 그 이후는 모두 무시 (export 시 빈 행으로 구분한 요약 섹션 등)
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break

        def cell(h: str):
            i = idx_of[h]
            return row[i] if i < len(row) else None

        raw_code = cell("종목코드")
        raw_name = cell("종목명")
        raw_avg = cell("평단가")
        raw_qty = cell("수량")
        raw_market = cell("시장") if "시장" in idx_of else None
        raw_currency = cell("통화") if "통화" in idx_of else None
        raw_buy_rate = cell("매수환율") if "매수환율" in idx_of else None

        # 종목코드: 숫자로 읽혔어도 문자열로 정규화 후 시장별로 검증 + 대문자
        if raw_code is None or str(raw_code).strip() == "":
            errors.append(f"{row_num}행: 종목코드가 비어 있습니다.")
            continue
        code = str(raw_code).strip().upper()
        # 엑셀이 숫자로 인식해 선행 0 손실된 경우 6자리로 패딩 (전부 숫자일 때만)
        if code.isdigit() and len(code) < 6:
            code = code.zfill(6)
        market = _normalize_excel_market(raw_market, code)
        currency = _normalize_excel_currency(raw_currency, market)
        if market == MARKET_KR and (len(code) != 6 or not code.isalnum()):
            errors.append(f"{row_num}행: 한국 종목코드 '{code}' 가 6자리 영숫자가 아닙니다.")
            continue
        if market == MARKET_US and not _US_TICKER_RE.match(code):
            errors.append(f"{row_num}행: 미국 종목코드 '{code}' 가 올바른 티커 형식이 아닙니다.")
            continue
        if code in seen_codes:
            errors.append(f"{row_num}행: 종목코드 '{code}' 가 중복되었습니다.")
            continue

        # 평단가/수량 변환. 수량은 소수점 3자리까지 허용한다.
        try:
            avg_price = float(raw_avg) if raw_avg is not None and str(raw_avg).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 평단가 '{raw_avg}' 가 숫자가 아닙니다.")
            continue
        try:
            quantity = round(float(raw_qty), 3) if raw_qty is not None and str(raw_qty).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 수량 '{raw_qty}' 가 숫자가 아닙니다.")
            continue
        if avg_price < 1:
            errors.append(f"{row_num}행: 평단가가 1 이상이어야 합니다.")
            continue
        if quantity <= 0:
            errors.append(f"{row_num}행: 수량이 0보다 커야 합니다.")
            continue

        name = str(raw_name).strip() if raw_name is not None and str(raw_name).strip() else code
        stock = {
            "code":      code,
            "name":      name,
            "market":    market,
            "currency":  currency,
            "avg_price": round(avg_price, 4) if market == MARKET_US else int(round(avg_price)),
            "quantity":  quantity,
        }
        if market == MARKET_US and raw_buy_rate is not None and str(raw_buy_rate).strip() != "":
            try:
                buy_exchange_rate = float(raw_buy_rate)
            except (TypeError, ValueError):
                errors.append(f"{row_num}행: 매수환율 '{raw_buy_rate}' 가 숫자가 아닙니다.")
                continue
            if buy_exchange_rate > 0:
                stock["buy_exchange_rate"] = buy_exchange_rate

        stocks.append(normalize_stock_schema(stock))
        seen_codes.add(code)

    if errors:
        # 너무 길지 않게 상위 10개만 보여줌
        head = "\n".join(errors[:10])
        more = f"\n... 외 {len(errors) - 10}건" if len(errors) > 10 else ""
        raise ValueError("다음 항목에서 오류가 발생했습니다:\n\n" + head + more)

    if not stocks:
        raise ValueError("가져올 종목이 없습니다. (데이터 행을 찾지 못했습니다)")

    return normalize_stocks_schema(stocks)
