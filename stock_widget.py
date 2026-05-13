#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
한국 주식 위젯 v1.0
- 종목별 실시간 현재가 표시 (네이버 금융 API)
- ▼ 버튼 클릭 시 평단가·수량·손익 확장, 5초 후 자동 축소
- 시스템 트레이에서 종목 추가/제거
- 위치는 자동 저장
"""

import sys
import json
import os
import copy
import shutil
import requests
from datetime import datetime, timedelta

# ─── 공용 HTTP 세션 (TCP/TLS 연결 재사용으로 호출당 100~300ms 절감) ─────────
_SESSION = requests.Session()
_SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
})

from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton,
    QVBoxLayout, QHBoxLayout, QGridLayout, QDialog, QFormLayout,
    QLineEdit, QSpinBox, QDialogButtonBox,
    QSystemTrayIcon, QMenu, QFrame, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QStyle, QStyleOptionSpinBox, QStyledItemDelegate,
    QFileDialog, QRadioButton, QButtonGroup
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QPoint, QPointF, QRectF
from PyQt6.QtGui import QFont, QFontMetrics, QColor, QPixmap, QPainter, QPainterPath, QIcon, QAction, QBrush, QPen, QPolygon

# ─── 설정 파일 경로 ────────────────────────────────────────────────────────────
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stocks.json")
BACKUP_FILE = CONFIG_FILE + ".bak"

# ─── Excel import/export 컬럼 정의 ────────────────────────────────────────────
# 헤더 ↔ stocks.json 필드 매핑. 순서는 export 시 컬럼 순서가 됨.
EXCEL_COLUMNS = [
    ("종목코드", "code"),
    ("종목명",   "name"),
    ("평단가",   "avg_price"),
    ("수량",     "quantity"),
]

# ─── 색상 테마 (다크 / Catppuccin Mocha 계열) ────────────────────────────────
C = {
    "bg":       "#1e1e2e",
    "bg2":      "#181825",
    "surface":  "#313244",
    "surface2": "#45475a",
    "text":     "#cdd6f4",
    "subtext":  "#a6adc8",
    "blue":     "#89b4fa",
    "red":      "#f38ba8",
    "green":    "#a6e3a1",
    "border":   "#313244",
}

TRAY_MENU_STYLE = f"""
QMenu {{
    background: {C['bg']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 8px;
    padding: 4px;
}}
QMenu::item {{
    padding: 7px 20px;
    border-radius: 5px;
}}
QMenu::item:selected {{
    background: {C['surface']};
}}
QMenu::separator {{
    height: 1px;
    background: {C['border']};
    margin: 4px 8px;
}}
"""

DIALOG_STYLE = f"""
QDialog {{
    background: {C['bg']};
    color: {C['text']};
}}
QLabel {{
    color: {C['subtext']};
    font-size: 12px;
}}
QLineEdit, QSpinBox {{
    background: {C['surface']};
    color: {C['text']};
    border: 1px solid {C['surface2']};
    border-radius: 7px;
    padding: 7px 10px;
    font-size: 13px;
    selection-background-color: {C['blue']};
}}
QLineEdit:focus, QSpinBox:focus {{
    border: 1px solid {C['blue']};
}}
QSpinBox::up-button, QSpinBox::down-button {{
    background: {C['surface2']};
    border: none;
    width: 22px;
}}
QSpinBox::up-button {{ border-top-right-radius: 6px; }}
QSpinBox::down-button {{ border-bottom-right-radius: 6px; }}
QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
    background: {C['blue']};
}}
/* 화살표는 ArrowSpinBox.paintEvent에서 직접 그림 — Qt 기본 화살표는 숨김 */
QSpinBox::up-arrow, QSpinBox::down-arrow {{
    image: none;
    width: 0;
    height: 0;
}}
QPushButton {{
    background: {C['blue']};
    color: {C['bg']};
    border: none;
    border-radius: 7px;
    padding: 8px 20px;
    font-size: 13px;
    font-weight: bold;
}}
QPushButton:hover {{
    background: #b4befe;
}}
QPushButton[flat="true"] {{
    background: {C['surface']};
    color: {C['text']};
}}
QPushButton[flat="true"]:hover {{
    background: {C['surface2']};
}}
QTableWidget {{
    background: {C['bg2']};
    color: {C['text']};
    border: 1px solid {C['surface2']};
    border-radius: 7px;
    gridline-color: {C['surface']};
    selection-background-color: {C['surface2']};
    selection-color: {C['text']};
    font-size: 12px;
}}
QTableWidget::item {{
    padding: 6px 8px;
    border: none;
}}
QHeaderView::section {{
    background: {C['surface']};
    color: {C['subtext']};
    border: none;
    border-right: 1px solid {C['bg2']};
    padding: 6px 8px;
    font-size: 11px;
    font-weight: bold;
}}
QHeaderView::section:last {{
    border-right: none;
}}
"""


# ─── 네이버 금융 API ───────────────────────────────────────────────────────────
def fetch_stock(code: str) -> dict | None:
    """네이버 금융 모바일 API로 현재가 조회"""
    url = f"https://m.stock.naver.com/api/stock/{code}/basic"
    try:
        r = _SESSION.get(url, timeout=3)
        if r.status_code != 200:
            return None
        d = r.json()
        return {
            "name":         d.get("stockName", code),
            "price":        int(str(d.get("closePrice", "0")).replace(",", "")),
            "change_rate":  float(d.get("fluctuationsRatio", 0)),
            "change_price": int(str(d.get("compareToPreviousClosePrice", "0")).replace(",", "")),
        }
    except Exception as e:
        print(f"[fetch_stock] {code} 오류: {e}")
        return None


# ─── 네이버 금융 분봉 차트 API ───────────────────────────────────────────────
def fetch_minute_chart(code: str) -> dict | None:
    """네이버 금융 분봉 API로 당일 1분봉 시계열 조회.
    반환: {'prices': [float, ...], 'open': float} or None"""
    url = f"https://api.stock.naver.com/chart/domestic/item/{code}/minute"
    try:
        r = _SESSION.get(url, timeout=3)
        if r.status_code != 200:
            return None
        data = r.json()
        if not data:
            return None
        return {
            "prices": [float(d["currentPrice"]) for d in data],
            "open":   float(data[0]["openPrice"]),
        }
    except Exception as e:
        print(f"[fetch_minute_chart] {code} 오류: {e}")
        return None


# ─── 네이버 금융 일봉 차트 API (장 외 시간 폴백용) ──────────────────────────
def fetch_daily_chart(code: str, days: int = 45, max_candles: int = 30) -> dict | None:
    """최근 N 캘린더일 일봉 OHLC 시계열 조회.
    분봉이 비어있는 장 외 시간/주말/공휴일에 캔들 차트로 표시할 용도.
    반환: {'candles': [{'open','high','low','close'}, ...]} or None"""
    end = datetime.now()
    start = end - timedelta(days=days)
    url = (
        f"https://api.stock.naver.com/chart/domestic/item/{code}/day"
        f"?startDateTime={start.strftime('%Y%m%d')}"
        f"&endDateTime={end.strftime('%Y%m%d')}"
    )
    try:
        r = _SESSION.get(url, timeout=3)
        if r.status_code != 200:
            return None
        data = r.json()
        if not data:
            return None
        candles = [
            {
                "open":  float(d["openPrice"]),
                "high":  float(d["highPrice"]),
                "low":   float(d["lowPrice"]),
                "close": float(d["closePrice"]),
            }
            for d in data
        ]
        if max_candles > 0:
            candles = candles[-max_candles:]
        return {"candles": candles}
    except Exception as e:
        print(f"[fetch_daily_chart] {code} 오류: {e}")
        return None


# ─── Excel import/export ─────────────────────────────────────────────────────
def export_stocks_to_excel(stocks: list[dict], path: str,
                           current_prices: dict | None = None) -> None:
    """보유 종목을 .xlsx 로 내보내기.
    - 종목코드는 텍스트 셀로 저장 (선행 0 보존: '005930', '0183J0').
    - 위젯 위치(pos)는 제외 — 다른 PC에서는 화면 좌표가 달라 의미가 없음.
    - current_prices ({code: price}) 가 주어지면 시트 하단에 포트폴리오 요약
      (총 매입금액 / 평가금액 / 평가손익 / 수익률) 을 빈 행 한 줄로 분리해서 추가.
      import 시에는 빈 행 이후의 행을 모두 무시하므로 라운드트립에 영향 없음."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "보유종목"

    headers = [h for h, _ in EXCEL_COLUMNS]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for s in stocks:
        row = []
        for _, key in EXCEL_COLUMNS:
            if key == "code":
                row.append(str(s.get("code", "")))
            elif key == "name":
                row.append(s.get("name", s.get("code", "")))
            else:
                row.append(int(s.get(key, 0)))
        ws.append(row)

    # 종목코드 컬럼을 텍스트 포맷으로 (선행 0/영문 안전)
    code_col_idx = next(i for i, (_, k) in enumerate(EXCEL_COLUMNS, 1) if k == "code")
    code_letter = ws.cell(row=1, column=code_col_idx).column_letter
    for cell in ws[code_letter][1:]:   # 헤더 제외
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    # 컬럼 너비 자동 조정 (간단히 헤더+여유)
    widths = {"종목코드": 12, "종목명": 28, "평단가": 12, "수량": 10}
    for col_idx, (header, _) in enumerate(EXCEL_COLUMNS, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = widths.get(header, 14)

    # ── 포트폴리오 요약 (종목이 1개 이상일 때) ────────────────────────
    # 종목 표와 빈 행 한 줄로 분리. import 측에서 빈 행 이후를 모두 무시하므로
    # 라운드트립 안전.
    if stocks:
        total_invest = sum(
            int(s.get("avg_price", 0)) * int(s.get("quantity", 0)) for s in stocks
        )
        cp = current_prices or {}
        total_eval = 0
        for s in stocks:
            avg = int(s.get("avg_price", 0))
            qty = int(s.get("quantity", 0))
            # 현재가가 없는 종목은 평단가로 폴백 (평가손익 0 으로 계산됨)
            price = int(cp.get(s.get("code")) or avg)
            total_eval += price * qty
        profit = total_eval - total_invest
        prate  = (profit / total_invest * 100.0) if total_invest else 0.0

        # 빈 행 한 줄 띄우고 다음 행에 요약 헤더
        header_row = ws.max_row + 2
        ws.cell(row=header_row, column=1, value="포트폴리오 요약").font = bold

        rows = [
            ("총 매입금액", total_invest, "#,##0"),
            ("평가금액",   total_eval,   "#,##0"),
            ("평가손익",   profit,        "#,##0"),
            ("수익률 (%)", round(prate, 2), "0.00"),
        ]
        for i, (label, val, fmt) in enumerate(rows, 1):
            r = header_row + i
            ws.cell(row=r, column=1, value=label)
            val_cell = ws.cell(row=r, column=2, value=val)
            val_cell.number_format = fmt
            val_cell.alignment = Alignment(horizontal="right")

    wb.save(path)


def import_stocks_from_excel(path: str) -> list[dict]:
    """Excel 파일에서 보유 종목을 읽어 stocks.json 형식 dict 리스트로 반환.
    검증 실패 시 ValueError 를 발생시킨다 (메시지는 사용자에게 그대로 표시 가능)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        raise ValueError("시트가 비어 있습니다.")

    # 1행 헤더 읽기 (공백/None 안전)
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = [h for h, _ in EXCEL_COLUMNS]
    missing = [h for h in required if h not in header_row]
    if missing:
        raise ValueError(
            "필수 컬럼이 누락되었습니다: " + ", ".join(missing)
            + f"\n(필요한 헤더: {', '.join(required)})"
        )

    # 헤더명 → 컬럼 인덱스
    idx_of = {h: header_row.index(h) for h in required}

    stocks: list[dict] = []
    seen_codes: set[str] = set()
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 빈 행을 만나면 그 이후는 모두 무시 (export 시 빈 행으로 구분한 요약 섹션 등)
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break

        def cell(h: str):
            i = idx_of[h]
            return row[i] if i < len(row) else None

        raw_code = cell("종목코드")
        raw_name = cell("종목명")
        raw_avg  = cell("평단가")
        raw_qty  = cell("수량")

        # 종목코드: 숫자로 읽혔어도 문자열로 정규화 후 6자 영숫자 검증 + 대문자
        if raw_code is None or str(raw_code).strip() == "":
            errors.append(f"{row_num}행: 종목코드가 비어 있습니다.")
            continue
        code = str(raw_code).strip().upper()
        # 엑셀이 숫자로 인식해 선행 0 손실된 경우 6자리로 패딩 (전부 숫자일 때만)
        if code.isdigit() and len(code) < 6:
            code = code.zfill(6)
        if len(code) != 6 or not code.isalnum():
            errors.append(f"{row_num}행: 종목코드 '{code}' 가 6자리 영숫자가 아닙니다.")
            continue
        if code in seen_codes:
            errors.append(f"{row_num}행: 종목코드 '{code}' 가 중복되었습니다.")
            continue

        # 평단가/수량: 정수 변환
        try:
            avg_price = int(float(raw_avg)) if raw_avg is not None and str(raw_avg).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 평단가 '{raw_avg}' 가 숫자가 아닙니다.")
            continue
        try:
            quantity = int(float(raw_qty)) if raw_qty is not None and str(raw_qty).strip() != "" else 0
        except (TypeError, ValueError):
            errors.append(f"{row_num}행: 수량 '{raw_qty}' 가 숫자가 아닙니다.")
            continue
        if avg_price < 1:
            errors.append(f"{row_num}행: 평단가가 1 이상이어야 합니다.")
            continue
        if quantity < 1:
            errors.append(f"{row_num}행: 수량이 1 이상이어야 합니다.")
            continue

        name = str(raw_name).strip() if raw_name is not None and str(raw_name).strip() else code

        stocks.append({
            "code":      code,
            "name":      name,
            "avg_price": avg_price,
            "quantity":  quantity,
        })
        seen_codes.add(code)

    if errors:
        # 너무 길지 않게 상위 10개만 보여줌
        head = "\n".join(errors[:10])
        more = f"\n... 외 {len(errors) - 10}건" if len(errors) > 10 else ""
        raise ValueError("다음 항목에서 오류가 발생했습니다:\n\n" + head + more)

    if not stocks:
        raise ValueError("가져올 종목이 없습니다. (데이터 행을 찾지 못했습니다)")

    return stocks


# ─── Excel import 모드 선택 다이얼로그 ───────────────────────────────────────
class ImportModeDialog(QDialog):
    """덮어쓰기 / 병합 모드 선택. accept 시 self.mode 에 'overwrite' 또는 'merge'."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("가져오기 모드")
        self.setFixedSize(360, 220)
        self.setStyleSheet(DIALOG_STYLE)
        self.mode: str = "merge"

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 18)
        root.setSpacing(10)

        title = QLabel("가져오기 방식을 선택하세요")
        title.setStyleSheet(f"color: {C['text']}; font-size: 13px; font-weight: bold;")
        root.addWidget(title)

        desc = QLabel(
            "기존 stocks.json 은 자동으로 stocks.json.bak 에 백업됩니다."
        )
        desc.setWordWrap(True)
        desc.setStyleSheet(f"color: {C['subtext']}; font-size: 11px;")
        root.addWidget(desc)

        root.addSpacing(4)

        radio_style = (
            f"QRadioButton {{ color: {C['text']}; font-size: 12px; padding: 4px 0; }}"
            f"QRadioButton::indicator {{ width: 14px; height: 14px; }}"
        )

        self.merge_rb = QRadioButton("병합 — 같은 종목코드는 Excel 값으로 갱신, 나머지는 유지")
        self.overwrite_rb = QRadioButton("덮어쓰기 — 기존 종목을 모두 삭제하고 Excel 내용으로 교체")
        self.merge_rb.setStyleSheet(radio_style)
        self.overwrite_rb.setStyleSheet(radio_style)
        self.merge_rb.setChecked(True)

        group = QButtonGroup(self)
        group.addButton(self.merge_rb)
        group.addButton(self.overwrite_rb)

        root.addWidget(self.merge_rb)
        root.addWidget(self.overwrite_rb)
        root.addStretch()

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("가져오기")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("취소")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setProperty("flat", "true")
        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def _on_ok(self):
        self.mode = "overwrite" if self.overwrite_rb.isChecked() else "merge"
        self.accept()


# ─── 가격 미니 차트 (sparkline) ───────────────────────────────────────────────
class SparklineWidget(QWidget):
    """미니 가격 차트. 두 가지 모드 지원.
    - line  : 당일 1분봉 라인 + area, 시초가 대비 색상 결정, 전일 종가 점선
    - candle: 최근 N일 일봉 캔들 (양봉=빨강, 음봉=파랑)"""

    W = 100   # 차트 너비
    H = 40    # 차트 높이

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(self.W, self.H)
        self.mode: str = "line"
        self.prices: list[float] = []
        self.open_price: float = 0.0
        self.prev_close: float = 0.0   # 전일 종가 (가로 점선 표시용, line 모드 전용)
        self.candles: list[dict] = []  # OHLC dict 리스트 (candle 모드)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

    def set_data(self, prices: list[float], open_price: float, prev_close: float = 0.0):
        self.mode = "line"
        self.prices = prices
        self.open_price = open_price
        self.prev_close = prev_close
        self.update()

    def set_candles(self, candles: list[dict]):
        self.mode = "candle"
        self.candles = candles
        self.update()

    def paintEvent(self, event):
        if self.mode == "candle":
            self._paint_candles()
        else:
            self._paint_line()

    # ── 라인 모드 (당일 분봉) ────────────────────────────────────────────
    def _paint_line(self):
        prices = self.prices
        if not prices or len(prices) < 2:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        last = prices[-1]
        op = self.open_price or prices[0]
        color = QColor(C['red']) if last >= op else QColor(C['blue'])
        fill = QColor(color)
        fill.setAlpha(50)

        pad = 4
        w = self.W - 2 * pad
        h = self.H - 2 * pad

        # y범위: 가격뿐 아니라 전일 종가도 포함시켜 점선이 항상 영역 안에 들어오게
        y_values = list(prices)
        if self.prev_close > 0:
            y_values.append(self.prev_close)
        mn = min(y_values)
        mx = max(y_values)
        rng = (mx - mn) if mx > mn else 1.0

        def y_of(price: float) -> float:
            return pad + (1 - (price - mn) / rng) * h

        # x축은 거래시간 전체(09:00~15:30, 약 391분봉) 기준으로 절대 매핑.
        # 단, 장 초반에 너무 좁아 보이지 않도록 최소 가시 영역(15%) 보장.
        TOTAL_BARS = 391
        MIN_VISIBLE_RATIO = 0.15
        actual_ratio = (len(prices) - 1) / (TOTAL_BARS - 1)
        visible_ratio = min(max(actual_ratio, MIN_VISIBLE_RATIO), 1.0)

        n = len(prices)
        pts: list[QPointF] = []
        for i, p in enumerate(prices):
            x = pad + (i / (n - 1)) * visible_ratio * w
            pts.append(QPointF(x, y_of(p)))

        # area fill (라인 아래 반투명 채움)
        area = QPainterPath()
        area.moveTo(pts[0])
        for pt in pts[1:]:
            area.lineTo(pt)
        area.lineTo(pts[-1].x(), pad + h)
        area.lineTo(pts[0].x(), pad + h)
        area.closeSubpath()
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(fill))
        painter.drawPath(area)

        # 전일 종가 기준선 (가로 점선) — 촘촘한 dot, 살짝 흐린 색
        if self.prev_close > 0:
            line_y = y_of(self.prev_close)
            pen_color = QColor(C['subtext'])
            pen_color.setAlpha(180)         # 살짝 흐리게
            dotted = QPen(pen_color)
            dotted.setWidthF(0.8)            # 얇게
            dotted.setDashPattern([1, 2])    # 1px on, 2px off (거의 dot)
            painter.setPen(dotted)
            painter.drawLine(QPointF(pad, line_y), QPointF(pad + w, line_y))

        # line stroke (현재가 라인)
        line = QPainterPath()
        line.moveTo(pts[0])
        for pt in pts[1:]:
            line.lineTo(pt)
        painter.setPen(QPen(color, 1.3))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawPath(line)

        painter.end()

    # ── 캔들 모드 (일봉) ────────────────────────────────────────────────
    def _paint_candles(self):
        candles = self.candles
        if not candles:
            return

        painter = QPainter(self)
        # 캔들은 픽셀 정렬이 더 선명 — antialias off
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, False)

        pad = 3
        w = self.W - 2 * pad
        h = self.H - 2 * pad

        mn = min(c["low"]  for c in candles)
        mx = max(c["high"] for c in candles)
        rng = (mx - mn) if mx > mn else 1.0

        def y_of(price: float) -> float:
            return pad + (1 - (price - mn) / rng) * h

        n = len(candles)
        slot = w / n
        body_w = max(1.5, slot * 0.7)

        red  = QColor(C['red'])
        blue = QColor(C['blue'])

        for i, c in enumerate(candles):
            cx = pad + (i + 0.5) * slot
            up = c["close"] >= c["open"]
            color = red if up else blue

            # 심지(고가–저가)
            painter.setPen(QPen(color, 0.8))
            painter.drawLine(
                QPointF(cx, y_of(c["high"])),
                QPointF(cx, y_of(c["low"])),
            )

            # 몸통(시가↔종가)
            y_open  = y_of(c["open"])
            y_close = y_of(c["close"])
            body_top = min(y_open, y_close)
            body_h   = max(1.0, abs(y_close - y_open))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QBrush(color))
            painter.drawRect(QRectF(cx - body_w / 2, body_top, body_w, body_h))

        painter.end()


# ─── 포커스 진입 시 자동 전체선택 ────────────────────────────────────────────
class _SelectAllOnFocus:
    """Mixin: 포커스가 들어오면 내용을 자동으로 전체 선택.
    selectAll() 메서드가 있는 위젯(QLineEdit·QSpinBox 등)과 혼합해 사용.

    focusInEvent 직후 Qt 내부에서 selection이 해제될 수 있어
    QTimer.singleShot(0, ...)으로 다음 이벤트 루프 tick에 호출한다."""

    def focusInEvent(self, event):
        super().focusInEvent(event)
        QTimer.singleShot(0, self.selectAll)


class AutoSelectLineEdit(_SelectAllOnFocus, QLineEdit):
    pass


class AutoSelectSpinBox(_SelectAllOnFocus, QSpinBox):
    pass


# ─── 화살표를 직접 그리는 QSpinBox ───────────────────────────────────────────
class ArrowSpinBox(AutoSelectSpinBox):
    """다크 stylesheet 환경에서 ▲▼ 화살표를 paintEvent로 직접 그림.
    PyQt6의 ::up-arrow / ::down-arrow가 CSS triangle·inline SVG 모두
    안 먹는 이슈를 회피한다. 포커스 시 자동 전체선택은 부모(AutoSelectSpinBox)
    에서 처리."""

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.buttonSymbols() == QSpinBox.ButtonSymbols.NoButtons:
            return

        # 정확한 up/down 버튼 영역 얻기
        opt = QStyleOptionSpinBox()
        self.initStyleOption(opt)
        style = self.style()
        up_rect = style.subControlRect(
            QStyle.ComplexControl.CC_SpinBox, opt,
            QStyle.SubControl.SC_SpinBoxUp, self)
        down_rect = style.subControlRect(
            QStyle.ComplexControl.CC_SpinBox, opt,
            QStyle.SubControl.SC_SpinBoxDown, self)

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QBrush(QColor(C['text'])))
        painter.setPen(Qt.PenStyle.NoPen)

        # 위 ▲
        cx, cy = up_rect.center().x(), up_rect.center().y()
        painter.drawPolygon(QPolygon([
            QPoint(cx,     cy - 3),
            QPoint(cx - 4, cy + 2),
            QPoint(cx + 4, cy + 2),
        ]))
        # 아래 ▼
        cx, cy = down_rect.center().x(), down_rect.center().y()
        painter.drawPolygon(QPolygon([
            QPoint(cx - 4, cy - 2),
            QPoint(cx + 4, cy - 2),
            QPoint(cx,     cy + 3),
        ]))
        painter.end()


# ─── 종목 추가 / 수정 다이얼로그 ──────────────────────────────────────────────
class StockDialog(QDialog):
    def __init__(self, parent=None, data: dict | None = None):
        super().__init__(parent)
        self.is_edit = data is not None
        self.setWindowTitle("종목 수정" if self.is_edit else "종목 추가")
        self.setFixedSize(340, 270)
        self.setStyleSheet(DIALOG_STYLE)

        layout = QFormLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(24, 24, 24, 20)
        # 라벨과 입력 위젯의 세로 중심을 일치시킴 (이슈 #2)
        layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        # 종목코드 (포커스 시 자동 전체선택)
        self.code_edit = AutoSelectLineEdit()
        self.code_edit.setPlaceholderText("예: 005930  (삼성전자)")
        self.code_edit.editingFinished.connect(self._preview_name)
        layout.addRow(self._row_label("종목 코드"), self.code_edit)

        # 종목명 미리보기 (코드 입력 후 자동 조회, 이슈 #2)
        self.preview_lbl = QLabel("─")
        self._set_preview_neutral()
        layout.addRow(self._row_label("종목명"), self.preview_lbl)

        # 평단가 (화살표 버튼 제거 + 포커스 시 자동 전체선택)
        self.avg_spin = AutoSelectSpinBox()
        self.avg_spin.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
        self.avg_spin.setRange(1, 10_000_000)
        self.avg_spin.setSingleStep(100)
        self.avg_spin.setSuffix("  원")
        layout.addRow(self._row_label("평단가"), self.avg_spin)

        # 수량 (paintEvent로 ▲▼ 화살표 직접 그림)
        self.qty_spin = ArrowSpinBox()
        self.qty_spin.setRange(1, 1_000_000)
        self.qty_spin.setSuffix("  주")
        layout.addRow(self._row_label("수  량"), self.qty_spin)

        # 기존 데이터 채우기
        if self.is_edit:
            self.code_edit.setText(data["code"])
            self.code_edit.setReadOnly(True)
            self.avg_spin.setValue(int(data.get("avg_price", 0)))
            self.qty_spin.setValue(int(data.get("quantity", 1)))
            if data.get("name"):
                self._set_preview_found(data["name"])

        # 버튼
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("확인")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("취소")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setProperty("flat", "true")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    # ── 라벨 생성기 (입력 위젯과 세로 중심 정렬, 이슈 #2) ────────────────
    @staticmethod
    def _row_label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        lbl.setMinimumHeight(34)   # QLineEdit/QSpinBox 높이와 매칭
        return lbl

    # ── 종목명 자동 미리보기 ─────────────────────────────────────────────
    def _preview_name(self):
        code = self.code_edit.text().strip().upper()
        if not code:
            self._set_preview_neutral()
            return
        if len(code) != 6 or not code.isalnum():
            self._set_preview_hint("6자리 코드를 입력하세요 (숫자/영문)")
            return
        self._set_preview_hint("조회 중...")
        self.preview_lbl.repaint()
        result = fetch_stock(code)
        if result:
            self._set_preview_found(result["name"])
        else:
            self._set_preview_error("찾을 수 없는 종목")

    def _set_preview_neutral(self):
        self.preview_lbl.setText("─")
        self.preview_lbl.setStyleSheet(
            f"color: {C['subtext']}; font-size: 12px; padding-left: 4px;"
        )

    def _set_preview_hint(self, msg: str):
        self.preview_lbl.setText(msg)
        self.preview_lbl.setStyleSheet(
            f"color: {C['subtext']}; font-size: 12px; font-style: italic; padding-left: 4px;"
        )

    def _set_preview_found(self, name: str):
        self.preview_lbl.setText(name)
        self.preview_lbl.setStyleSheet(
            f"color: {C['text']}; font-size: 13px; font-weight: bold; padding-left: 4px;"
        )

    def _set_preview_error(self, msg: str):
        self.preview_lbl.setText(msg)
        self.preview_lbl.setStyleSheet(
            f"color: {C['red']}; font-size: 12px; font-style: italic; padding-left: 4px;"
        )

    def get_data(self) -> dict:
        return {
            "code":      self.code_edit.text().strip().upper(),
            "avg_price": self.avg_spin.value(),
            "quantity":  self.qty_spin.value(),
        }


# ─── 좁은 셀에서도 입력값이 잘리지 않도록 editor 폭을 약간만 늘리는 delegate ──
class WideEditorDelegate(QStyledItemDelegate):
    """편집 진입 시 editor 가로폭을 셀 폭 + PADDING 으로 임시 확장.
    셀 자체 너비는 그대로, editor 만 약간 넓어져 cursor·입력값이 잘리지 않게."""

    PADDING = 15   # 셀 폭에 추가할 여유 (cursor + 한두 자 입력 공간)

    def updateEditorGeometry(self, editor, option, index):
        rect = option.rect
        new_w = rect.width() + self.PADDING
        editor.setGeometry(rect.x(), rect.y(), new_w, rect.height())


# ─── 종목 일괄 관리 다이얼로그 ────────────────────────────────────────────────
class ManageStocksDialog(QDialog):
    """현재 보유 종목들을 표 형태로 일괄 관리하는 다이얼로그."""

    COLS = ["종목명", "종목코드", "평단가", "수량", "평가손익"]

    def __init__(self, stocks: list[dict], current_prices: dict | None = None, parent=None):
        super().__init__(parent)
        self._stocks: list[dict] = stocks   # 호출측에서 deepcopy 해서 전달
        self._current_prices: dict = current_prices or {}   # {code: 현재가}
        self._suppress_change: bool = False   # itemChanged 재귀 차단용

        self.setWindowTitle("종목 관리")
        self.setMinimumSize(640, 400)
        self.setStyleSheet(DIALOG_STYLE)

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 16)
        root.setSpacing(12)

        # ── 표 ─────────────────────────────────────────────────────────────
        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.verticalHeader().setVisible(False)
        # 더블클릭/EditKey/AnyKey 로 셀 인라인 편집 진입 (편집 가능 셀은 _fill_row 에서 지정)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setShowGrid(False)
        self.table.setAlternatingRowColors(False)

        # 드래그로 행 순서 변경
        self.table.setDragEnabled(True)
        self.table.setAcceptDrops(True)
        self.table.viewport().setAcceptDrops(True)
        self.table.setDropIndicatorShown(True)
        self.table.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.table.setDragDropOverwriteMode(False)

        # 컬럼 너비 정책
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)         # 종목명
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setStretchLastSection(False)

        # 헤더 클릭 자동 정렬은 사용하지 않음 (명시적 "정렬" 버튼으로 대체)
        hdr.setSectionsClickable(False)
        hdr.setSortIndicatorShown(False)

        # 평단가/수량 인라인 편집 시 editor 폭을 키워서 입력값이 잘리지 않게
        self._wide_delegate = WideEditorDelegate(self)
        self.table.setItemDelegateForColumn(2, self._wide_delegate)
        self.table.setItemDelegateForColumn(3, self._wide_delegate)

        # 더블클릭: 평단가/수량 셀은 Qt 가 인라인 편집을 처리하므로 패스,
        # 그 외 셀에서는 기존처럼 종목 수정 팝업을 띄움
        self.table.doubleClicked.connect(self._on_double_clicked)

        # 인라인 편집 결과 반영
        self.table.itemChanged.connect(self._on_item_changed)

        # 드래그 정렬: 모델의 rowsMoved 시그널로 self._stocks 순서 동기화
        self.table.model().rowsMoved.connect(self._on_rows_moved)

        root.addWidget(self.table, 1)

        # ── 행 액션 버튼 (추가 / 수정 / 삭제) ─────────────────────────────
        action_row = QHBoxLayout()
        action_row.setSpacing(8)

        add_btn = QPushButton("➕  추가")
        add_btn.clicked.connect(self._add)
        action_row.addWidget(add_btn)

        edit_btn = QPushButton("✏  수정")
        edit_btn.setProperty("flat", "true")
        edit_btn.clicked.connect(self._edit_selected)
        action_row.addWidget(edit_btn)

        del_btn = QPushButton("🗑  삭제")
        del_btn.setProperty("flat", "true")
        del_btn.clicked.connect(self._delete_selected)
        action_row.addWidget(del_btn)

        action_row.addStretch()

        # 평가손익 내림차순 정렬 (명시적 버튼, 자동 정렬은 안 함)
        sort_btn = QPushButton("📊  평가손익 정렬")
        sort_btn.setProperty("flat", "true")
        sort_btn.clicked.connect(self._sort_by_profit_desc)
        action_row.addWidget(sort_btn)

        root.addLayout(action_row)

        # ── 확인 / 취소 ────────────────────────────────────────────────────
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("확인")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("취소")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setProperty("flat", "true")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self._rebuild_table()

    # ── 표 동기화 ─────────────────────────────────────────────────────────
    def _rebuild_table(self, select_row: int | None = None):
        """self._stocks 기준으로 표를 다시 그림."""
        # rowsMoved / itemChanged 신호가 재구성 중에 발화되지 않도록 일시 차단
        self.table.model().rowsMoved.disconnect(self._on_rows_moved)
        self._suppress_change = True
        try:
            self.table.setRowCount(0)
            for s in self._stocks:
                row = self.table.rowCount()
                self.table.insertRow(row)
                self._fill_row(row, s)
        finally:
            self._suppress_change = False
            self.table.model().rowsMoved.connect(self._on_rows_moved)

        if select_row is not None and 0 <= select_row < self.table.rowCount():
            self.table.selectRow(select_row)

    def _fill_row(self, row: int, s: dict):
        name  = s.get("name", s["code"])
        code  = s["code"]
        avg_p = int(s.get("avg_price", 0))
        qty_n = int(s.get("quantity", 0))
        avg   = f"{avg_p:,} 원"
        qty   = f"{qty_n:,} 주"

        # 평가손익 = (현재가 - 평단가) * 수량. 현재가 없으면 평단가 fallback → 0
        cur_p  = int(self._current_prices.get(code, avg_p))
        profit = (cur_p - avg_p) * qty_n
        if profit > 0:
            profit_text  = f"+{profit:,} 원"
            profit_color = C['red']    # 이익 = 빨강 (한국 컨벤션)
        elif profit < 0:
            profit_text  = f"{profit:,} 원"     # 음수면 자체 '-' 표시
            profit_color = C['blue']   # 손실 = 파랑
        else:
            profit_text  = "0 원"
            profit_color = None

        cells = [name, code, avg, qty, profit_text]
        for col, text in enumerate(cells):
            item = QTableWidgetItem(text)
            # 평단가/수량/평가손익은 우측 정렬
            if col in (2, 3, 4):
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            else:
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            # 평가손익 셀에 색상 적용
            if col == 4 and profit_color is not None:
                item.setForeground(QColor(profit_color))
                f = item.font()
                f.setBold(True)
                item.setFont(f)
            # 평단가(2)/수량(3) 셀은 인라인 편집 가능
            base_flags = (
                Qt.ItemFlag.ItemIsSelectable
                | Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsDragEnabled
            )
            if col in (2, 3):
                base_flags |= Qt.ItemFlag.ItemIsEditable
            item.setFlags(base_flags)
            self.table.setItem(row, col, item)

    # ── 더블클릭: 평단가/수량은 인라인 편집, 그 외는 종목 수정 팝업 ──────
    def _on_double_clicked(self, index):
        if index.column() in (2, 3):
            return   # Qt 의 자동 인라인 편집에 맡김
        self._edit_selected()

    # ── 인라인 편집 결과 반영 ────────────────────────────────────────────
    def _on_item_changed(self, item):
        if self._suppress_change or item is None:
            return
        row, col = item.row(), item.column()
        if col not in (2, 3) or row < 0 or row >= len(self._stocks):
            return

        # 사용자가 입력한 텍스트에서 숫자만 추출
        text = item.text().strip()
        digits = "".join(c for c in text if c.isdigit())
        s = self._stocks[row]

        if not digits or int(digits) <= 0:
            # 잘못된 입력 → 원래 값으로 복원
            self._suppress_change = True
            if col == 2:
                item.setText(f"{int(s.get('avg_price', 0)):,} 원")
            else:
                item.setText(f"{int(s.get('quantity', 0)):,} 주")
            self._suppress_change = False
            return

        value = int(digits)
        if col == 2:
            s["avg_price"] = value
            suffix = "원"
        else:
            s["quantity"] = value
            suffix = "주"

        # 표시 형식 (쉼표 + 단위) 재포맷
        self._suppress_change = True
        item.setText(f"{value:,} {suffix}")
        self._suppress_change = False

        # 평가손익 셀 즉시 갱신
        self._refresh_profit_cell(row)

    def _refresh_profit_cell(self, row: int):
        s = self._stocks[row]
        code = s["code"]
        avg = int(s.get("avg_price", 0))
        qty = int(s.get("quantity", 0))
        cur = int(self._current_prices.get(code, avg))
        profit = (cur - avg) * qty

        if profit > 0:
            text, color = f"+{profit:,} 원", C['red']
        elif profit < 0:
            text, color = f"{profit:,} 원", C['blue']
        else:
            text, color = "0 원", None

        item = self.table.item(row, 4)
        if item is None:
            return
        self._suppress_change = True
        item.setText(text)
        if color:
            item.setForeground(QColor(color))
            f = item.font()
            f.setBold(True)
            item.setFont(f)
        else:
            item.setForeground(QColor(C['text']))
            f = item.font()
            f.setBold(False)
            item.setFont(f)
        self._suppress_change = False

    # ── 평가손익 내림차순 정렬 (명시적 버튼) ─────────────────────────────
    def _sort_by_profit_desc(self):
        def key_for(s: dict):
            avg = int(s.get("avg_price", 0))
            qty = int(s.get("quantity", 0))
            cur = int(self._current_prices.get(s["code"], avg))
            return (cur - avg) * qty
        self._stocks.sort(key=key_for, reverse=True)
        self._rebuild_table()

    # ── 드래그 정렬 핸들러 ────────────────────────────────────────────────
    def _on_rows_moved(self, parent, start, end, dest_parent, dest_row):
        # 단일 행만 이동(SingleSelection) — 한 항목을 옮긴 결과를 self._stocks 에 반영
        # Qt 의 dest_row 는 "이동 전 좌표계" 기준이므로 보정 필요
        item = self._stocks.pop(start)
        insert_at = dest_row if dest_row < start else dest_row - 1
        insert_at = max(0, min(insert_at, len(self._stocks)))
        self._stocks.insert(insert_at, item)
        # 표는 Qt 가 이미 옮긴 상태이므로 재구성 불필요

    # ── 액션 ───────────────────────────────────────────────────────────────
    def _add(self):
        dlg = StockDialog(parent=self)
        if not dlg.exec():
            return
        d = dlg.get_data()
        code = d["code"]
        if not code:
            return
        if any(s["code"] == code for s in self._stocks):
            QMessageBox.information(self, "알림", f"'{code}'는 이미 추가되어 있습니다.")
            return

        result = fetch_stock(code)
        if not result:
            QMessageBox.warning(
                self, "조회 실패",
                f"종목코드 '{code}'를 찾을 수 없습니다.\n코드를 다시 확인해 주세요."
            )
            return

        d["name"] = result["name"]
        self._stocks.append(d)
        # 현재가도 캐시해 두면 평가손익이 즉시 계산됨
        self._current_prices[code] = int(result["price"])
        self._rebuild_table(select_row=len(self._stocks) - 1)

    def _edit_selected(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self._stocks):
            return
        dlg = StockDialog(parent=self, data=self._stocks[row])
        if not dlg.exec():
            return
        new = dlg.get_data()
        self._stocks[row]["avg_price"] = new["avg_price"]
        self._stocks[row]["quantity"]  = new["quantity"]
        self._rebuild_table(select_row=row)

    def _delete_selected(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self._stocks):
            return
        name = self._stocks[row].get("name", self._stocks[row]["code"])
        ret = QMessageBox.question(
            self, "삭제 확인",
            f"'{name}' 을(를) 삭제할까요?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if ret != QMessageBox.StandardButton.Yes:
            return
        self._stocks.pop(row)
        next_sel = min(row, len(self._stocks) - 1) if self._stocks else None
        self._rebuild_table(select_row=next_sel)

    def get_stocks(self) -> list[dict]:
        return self._stocks


# ─── 개별 주식 위젯 ───────────────────────────────────────────────────────────
class StockWidget(QWidget):
    """화면에 떠있는 하나의 주식 위젯"""

    deleted        = pyqtSignal(str)   # code 전달
    edited         = pyqtSignal(str)   # 수정 완료 후 저장 요청
    price_updated  = pyqtSignal(str)   # 현재가 갱신 시 (마스터 위젯 재집계용)

    MIN_W      = 240    # 기본(최소) 가로폭
    COMPACT_H  = 58     # 축소 높이 (2줄 레이아웃, 압축)
    EXPAND_H   = 214    # 확장 높이 (compact + 상세 패널 156)
    RADIUS     = 13     # 모서리 반지름

    def __init__(self, stock_data: dict, width: int | None = None, stagger_idx: int = 0):
        super().__init__()
        self.data = stock_data          # code, name, avg_price, quantity, pos
        self.current_price: int = 0
        self.is_expanded: bool = False
        self._drag_pos = None
        self._press_pos = None    # 좌클릭 시작 위치 (드래그/클릭 구분용)
        self._moved: bool = False # 일정 거리 이상 움직였는지
        self._stagger_idx = stagger_idx   # 동시 호출 분산용 인덱스

        # 외부에서 통일 너비를 받지 않으면 종목명 기준 자체 계산
        name = self.data.get("name", self.data["code"])
        self.W = width if width else self.calc_width_for_name(name)

        # 5초 자동 축소 타이머
        self.collapse_timer = QTimer(singleShot=True)
        self.collapse_timer.timeout.connect(self.collapse)

        # 가격은 5초마다, sparkline은 60초마다 갱신
        # (분봉 데이터는 1분 단위 생성이라 더 자주 호출해도 같은 데이터)
        self._prev_close: float = 0.0

        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self._fetch_price)

        self.chart_timer = QTimer()
        self.chart_timer.timeout.connect(self._fetch_chart)

        self._build_ui()

        # 타이머/첫 fetch를 stagger 인덱스만큼 지연시켜 시작.
        # 여러 위젯이 거의 같은 시점에 동시 HTTP 호출하지 않도록 분산.
        STAGGER_MS = 600   # 위젯당 약 0.6초 간격
        delay = self._stagger_idx * STAGGER_MS
        QTimer.singleShot(delay, self._start_fetching)

    def _start_fetching(self):
        """타이머 가동 + 즉시 1회 fetch (stagger 지연 후 호출)."""
        self.refresh_timer.start(5_000)
        self.chart_timer.start(60_000)
        self._fetch_price()
        self._fetch_chart()

    # ── 종목명에 맞춰 가로폭 계산 ─────────────────────────────────────────
    @staticmethod
    def calc_width_for_name(name: str) -> int:
        """종목명 픽셀 폭을 측정해 위젯 가로폭을 결정. 최소 MIN_W."""
        font = QFont("Malgun Gothic",8, QFont.Weight.Bold)
        fm = QFontMetrics(font)
        name_w = fm.horizontalAdvance(name)
        # 좌마진(14) + 정보~sparkline spacing(8) + sparkline(100) + 우마진(10) + 여유(6) = 138
        OVERHEAD = 138
        return max(StockWidget.MIN_W, name_w + OVERHEAD)

    # ── UI 구성 ────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(self.W, self.COMPACT_H)

        # ── 카드 배경 프레임
        self.card = QFrame(self)
        self.card.setObjectName("card")
        self.card.setGeometry(0, 0, self.W, self.COMPACT_H)
        self.card.setStyleSheet(f"""
            QFrame#card {{
                background: {C['bg']};
                border: 1px solid {C['border']};
                border-radius: {self.RADIUS}px;
            }}
        """)

        # ── 상단 compact 영역 (좌: 정보 / 우: 당일 sparkline) ──────────
        self.compact = QWidget(self.card)
        self.compact.setGeometry(0, 0, self.W, self.COMPACT_H)
        self.compact.setStyleSheet("background: transparent;")

        hl = QHBoxLayout(self.compact)
        hl.setContentsMargins(14, 5, 10, 5)
        hl.setSpacing(8)

        # 좌측: 종목명 + 가격 행
        info = QVBoxLayout()
        info.setContentsMargins(0, 0, 0, 0)
        info.setSpacing(1)

        # 1행: 종목명
        self.name_lbl = QLabel(self.data.get("name", self.data["code"]))
        self.name_lbl.setFont(QFont("Malgun Gothic",8, QFont.Weight.Bold))
        self.name_lbl.setStyleSheet(f"color: {C['subtext']};")
        info.addWidget(self.name_lbl)

        # 2행: 가격 + 등락률
        price_row = QHBoxLayout()
        price_row.setContentsMargins(0, 0, 0, 0)
        price_row.setSpacing(8)

        self.price_lbl = QLabel("─")
        self.price_lbl.setFont(QFont("Malgun Gothic",11, QFont.Weight.Bold))
        self.price_lbl.setStyleSheet(f"color: {C['text']};")
        price_row.addWidget(self.price_lbl)

        self.rate_lbl = QLabel("")
        self.rate_lbl.setFont(QFont("Malgun Gothic",9))
        self.rate_lbl.setStyleSheet(f"color: {C['subtext']};")
        price_row.addWidget(self.rate_lbl)
        price_row.addStretch()

        info.addLayout(price_row)
        hl.addLayout(info, 1)

        # 우측: 당일 sparkline 미니 차트
        self.sparkline = SparklineWidget(self.compact)
        hl.addWidget(self.sparkline, 0, Qt.AlignmentFlag.AlignVCenter)

        # ── 확장 패널 ────────────────────────────────────────────────────
        panel_h = self.EXPAND_H - self.COMPACT_H
        self.expand_panel = QWidget(self.card)
        self.expand_panel.setGeometry(0, self.COMPACT_H, self.W, panel_h)
        self.expand_panel.setStyleSheet("background: transparent;")
        self.expand_panel.hide()

        vl = QVBoxLayout(self.expand_panel)
        vl.setContentsMargins(14, 2, 14, 12)
        vl.setSpacing(2)

        # 구분선
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background: {C['border']}; max-height: 1px; border: none;")
        vl.addWidget(sep)
        vl.addSpacing(2)

        # 상세 행 생성
        self.avg_val    = self._make_row(vl, "평단가")
        self.qty_val    = self._make_row(vl, "보유수량")
        self.invest_val = self._make_row(vl, "투자원금")
        self.eval_val   = self._make_row(vl, "평가금액")

        # 손익 (강조)
        self.profit_val = self._make_row(vl, "평가손익", bold=True)
        self.prate_val  = self._make_row(vl, "수익률",   bold=True)

    # ── 외부에서 위젯 너비 변경 (통일 너비 적용용) ────────────────────
    def set_width(self, new_w: int):
        if new_w == self.W:
            return
        self.W = new_w
        cur_h = self.EXPAND_H if self.is_expanded else self.COMPACT_H
        self.setFixedWidth(new_w)
        self.card.setGeometry(0, 0, new_w, cur_h)
        self.compact.setGeometry(0, 0, new_w, self.COMPACT_H)
        panel_h = self.EXPAND_H - self.COMPACT_H
        self.expand_panel.setGeometry(0, self.COMPACT_H, new_w, panel_h)

    def _make_row(self, parent_layout, key_text: str, bold=False) -> QLabel:
        """키-값 한 줄 생성, 값 QLabel 반환"""
        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)

        key_lbl = QLabel(key_text)
        key_lbl.setStyleSheet(f"color: {C['subtext']}; font-size: 10px;")
        key_lbl.setFixedWidth(58)

        val_lbl = QLabel("─")
        style = f"color: {C['text']}; font-size: 11px;"
        if bold:
            style += " font-weight: bold;"
        val_lbl.setStyleSheet(style)
        val_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)

        row.addWidget(key_lbl)
        row.addWidget(val_lbl)
        parent_layout.addLayout(row)
        return val_lbl

    # ── 데이터 갱신 ────────────────────────────────────────────────────────
    def _fetch_price(self):
        """현재가/등락률 갱신 (5초 주기)."""
        result = fetch_stock(self.data["code"])
        if result:
            self.data["name"] = result["name"]
            self.name_lbl.setText(result["name"])
            self.current_price = result["price"]
            self._prev_close = float(result["price"] - result["change_price"])
            self._apply_price(result)
            self.price_updated.emit(self.data["code"])

    def _fetch_chart(self):
        """sparkline 갱신 (60초 주기) — 당일 분봉 우선, 비어있으면 최근 일봉 폴백."""
        chart = fetch_minute_chart(self.data["code"])
        if chart and len(chart["prices"]) >= 2:
            # 분봉 모드: 전일 종가 점선(=현재가 - 전일대비)도 함께 표시
            self.sparkline.set_data(chart["prices"], chart["open"], self._prev_close)
        else:
            # 일봉 모드: 최근 N일 캔들 차트로 폴백
            daily = fetch_daily_chart(self.data["code"])
            if daily:
                self.sparkline.set_candles(daily["candles"])

    def _apply_price(self, result: dict):
        price = result["price"]
        rate  = result["change_rate"]

        self.price_lbl.setText(f"{price:,}")

        if rate > 0:
            color = C["red"]
            sign  = "▲"
        elif rate < 0:
            color = C["blue"]
            sign  = "▼"
        else:
            color = C["subtext"]
            sign  = "  "

        self.price_lbl.setStyleSheet(f"color: {color}; font-size: 11px; font-weight: bold;")
        self.rate_lbl.setText(f"{sign}{abs(rate):.2f}%")
        self.rate_lbl.setStyleSheet(f"color: {color}; font-size: 9px;")

        self._update_detail(price)

    def _update_detail(self, price: int):
        avg    = self.data.get("avg_price", 0)
        qty    = self.data.get("quantity", 0)
        invest = avg * qty
        eval_  = price * qty
        profit = eval_ - invest
        prate  = (profit / invest * 100) if invest else 0

        sign  = "+" if profit >= 0 else ""
        color = C["red"] if profit >= 0 else C["blue"]

        self.avg_val.setText(f"{avg:,} 원")
        self.qty_val.setText(f"{qty:,} 주")
        self.invest_val.setText(f"{invest:,} 원")
        self.eval_val.setText(f"{eval_:,} 원")

        self.profit_val.setText(f"{sign}{profit:,} 원")
        self.profit_val.setStyleSheet(f"color: {color}; font-size: 11px; font-weight: bold;")
        self.prate_val.setText(f"{sign}{prate:.2f}%")
        self.prate_val.setStyleSheet(f"color: {color}; font-size: 11px; font-weight: bold;")

    # ── 확장 / 축소 ────────────────────────────────────────────────────────
    def toggle_expand(self):
        if self.is_expanded:
            self.collapse()
        else:
            self.expand()

    def expand(self):
        self.is_expanded = True
        self.expand_panel.show()
        self.setFixedHeight(self.EXPAND_H)
        self.card.setGeometry(0, 0, self.W, self.EXPAND_H)
        self.collapse_timer.start(5_000)   # 5초 뒤 자동 축소

    def collapse(self):
        self.is_expanded = False
        self.expand_panel.hide()
        self.setFixedHeight(self.COMPACT_H)
        self.card.setGeometry(0, 0, self.W, self.COMPACT_H)
        self.collapse_timer.stop()

    # ── 드래그 이동 + 클릭 토글 ──────────────────────────────────────────
    DRAG_THRESHOLD = 4   # 이 거리 이상 움직이면 드래그로 간주

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos  = event.globalPosition().toPoint() - self.pos()
            self._press_pos = event.globalPosition().toPoint()
            self._moved     = False

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.MouseButton.LeftButton and self._drag_pos:
            if not self._moved and self._press_pos:
                delta = event.globalPosition().toPoint() - self._press_pos
                if abs(delta.x()) > self.DRAG_THRESHOLD or abs(delta.y()) > self.DRAG_THRESHOLD:
                    self._moved = True
            if self._moved:
                self.move(event.globalPosition().toPoint() - self._drag_pos)

    def mouseReleaseEvent(self, event):
        # 드래그가 아니었으면(거의 안 움직임) = 클릭 → 확장/축소 토글
        if event.button() == Qt.MouseButton.LeftButton and not self._moved:
            self.toggle_expand()
        self._drag_pos  = None
        self._press_pos = None
        self._moved     = False

    # ── 우클릭 메뉴 ────────────────────────────────────────────────────────
    def contextMenuEvent(self, event):
        menu = QMenu(self)
        menu.setStyleSheet(TRAY_MENU_STYLE)
        edit_act = menu.addAction("✏️   수정")
        menu.addSeparator()
        del_act  = menu.addAction("🗑️   삭제")

        action = menu.exec(event.globalPos())
        if action == edit_act:
            self._open_edit()
        elif action == del_act:
            self.deleted.emit(self.data["code"])
            self.close()

    def _open_edit(self):
        dlg = StockDialog(data=self.data)
        if dlg.exec():
            new = dlg.get_data()
            self.data["avg_price"] = new["avg_price"]
            self.data["quantity"]  = new["quantity"]
            if self.current_price:
                self._update_detail(self.current_price)
            self.edited.emit(self.data["code"])


# ─── 포트폴리오 요약 마스터 위젯 ─────────────────────────────────────────────
class MasterWidget(QWidget):
    """포트폴리오 전체 요약을 표시하는 마스터 위젯.
    총 매입금액 / 평가금액 / 평가손익 / 수익률 4개 지표를 2×2 그리드로 표시.
    개별 종목 위젯과 동일한 다크 카드 스타일이며 드래그로 이동 가능."""

    H      = 96    # compact 카드 높이 (2×2 요약 그리드)
    RADIUS = 13
    DRAG_THRESHOLD = 4

    def __init__(self, width: int):
        super().__init__()
        # 가장 긴 종목명 기준 통일 폭과 동일하게 맞춤
        self.W = width
        self._drag_pos  = None
        self._press_pos = None
        self._moved     = False
        self.is_expanded: bool = False
        self.holdings: list[dict] = []   # [{"name", "profit", "profit_rate"}, ...]

        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(self.W, self.H)

        self.card = QFrame(self)
        self.card.setObjectName("master_card")
        self.card.setGeometry(0, 0, self.W, self.H)
        self.card.setStyleSheet(f"""
            QFrame#master_card {{
                background: {C['bg']};
                border: 1px solid {C['border']};
                border-radius: {self.RADIUS}px;
            }}
        """)

        # 상단 compact: 2x2 그리드 (제목 없음, 1행/2행 사이를 살짝 띄움)
        self.compact = QWidget(self.card)
        self.compact.setGeometry(0, 0, self.W, self.H)
        self.compact.setStyleSheet("background: transparent;")
        grid = QGridLayout(self.compact)
        grid.setContentsMargins(14, 12, 14, 12)
        grid.setHorizontalSpacing(20)
        grid.setVerticalSpacing(10)

        self.invest_val = self._make_cell(grid, 0, 0, "총 매입금액")
        self.eval_val   = self._make_cell(grid, 0, 1, "평가금액")
        self.profit_val = self._make_cell(grid, 1, 0, "평가손익", bold=True)
        self.prate_val  = self._make_cell(grid, 1, 1, "수익률",   bold=True)

        # 확장 패널 (클릭 시 종목별 손익 표시) — 초기 숨김
        self.expand_panel = QWidget(self.card)
        self.expand_panel.setStyleSheet("background: transparent;")
        self.expand_panel.hide()

    def _make_cell(self, grid: QGridLayout, row: int, col: int,
                   key_text: str, bold: bool = False) -> QLabel:
        cell = QVBoxLayout()
        cell.setContentsMargins(0, 0, 0, 0)
        cell.setSpacing(0)

        key_lbl = QLabel(key_text)
        key_lbl.setStyleSheet(f"color: {C['subtext']}; font-size: 10px;")
        cell.addWidget(key_lbl)

        style = f"color: {C['text']}; font-size: 13px;"
        if bold:
            style += " font-weight: bold;"
        val_lbl = QLabel("─")
        val_lbl.setStyleSheet(style)
        cell.addWidget(val_lbl)

        grid.addLayout(cell, row, col)
        return val_lbl

    # ── 외부에서 너비 변경 (개별 위젯 통일 폭에 맞춰 갱신) ───────────────
    def set_uniform_width(self, base_w: int):
        if base_w == self.W:
            return
        self.W = base_w
        self.setFixedWidth(base_w)
        cur_h = self.height()
        self.card.setGeometry(0, 0, base_w, cur_h)
        self.compact.setGeometry(0, 0, base_w, self.H)
        if self.is_expanded:
            panel_h = cur_h - self.H
            self.expand_panel.setGeometry(0, self.H, base_w, panel_h)

    # ── 지표 갱신 ────────────────────────────────────────────────────────
    def update_metrics(self, total_invest: int, total_eval: int):
        profit = total_eval - total_invest
        prate  = (profit / total_invest * 100.0) if total_invest else 0.0

        # 한국 시장 컨벤션과 일관: 이익=빨강, 손실=파랑
        if profit > 0:
            color = C['red']
            sign  = "+"
        elif profit < 0:
            color = C['blue']
            sign  = ""   # 음수면 자체적으로 '-' 가 붙음
        else:
            color = C['subtext']
            sign  = ""

        self.invest_val.setText(f"{total_invest:,} 원")
        self.eval_val.setText(f"{total_eval:,} 원")
        self.profit_val.setText(f"{sign}{profit:,} 원")
        self.profit_val.setStyleSheet(
            f"color: {color}; font-size: 13px; font-weight: bold;"
        )
        self.prate_val.setText(f"{sign}{prate:.2f}%")
        self.prate_val.setStyleSheet(
            f"color: {color}; font-size: 13px; font-weight: bold;"
        )

    def clear_metrics(self):
        """종목이 하나도 없을 때 0/빈 표시로 초기화."""
        self.invest_val.setText("0 원")
        self.eval_val.setText("0 원")
        self.profit_val.setText("─")
        self.profit_val.setStyleSheet(
            f"color: {C['subtext']}; font-size: 13px; font-weight: bold;"
        )
        self.prate_val.setText("─")
        self.prate_val.setStyleSheet(
            f"color: {C['subtext']}; font-size: 13px; font-weight: bold;"
        )
        self.holdings = []
        if self.is_expanded:
            self.collapse()

    # ── 보유 종목 목록 표시 ──────────────────────────────────────────────
    ROW_H        = 20    # 종목 1행 높이 (폰트 11 + 약간의 여유)
    ROW_SPACING  = 4
    PANEL_TOP    = 6
    PANEL_BOTTOM = 10

    def update_holdings(self, holdings: list[dict]):
        """holdings: [{"name": str, "profit": int, "profit_rate": float}, ...]
        펼친 상태면 즉시 다시 그리고 카드 높이도 재조정."""
        self.holdings = holdings
        if self.is_expanded:
            self._render_holdings()
            self._resize_to_expanded()

    def _calc_panel_height(self) -> int:
        n = len(self.holdings)
        if n == 0:
            return 0
        # 구분선(1px) + top/bottom padding + N행 + (N-1) row spacing
        return (
            self.PANEL_TOP + 1 + self.PANEL_TOP
            + n * self.ROW_H + max(0, n - 1) * self.ROW_SPACING
            + self.PANEL_BOTTOM
        )

    def _resize_to_expanded(self):
        panel_h = self._calc_panel_height()
        total_h = self.H + panel_h
        self.setFixedHeight(total_h)
        self.card.setGeometry(0, 0, self.W, total_h)
        self.expand_panel.setGeometry(0, self.H, self.W, panel_h)

    def _render_holdings(self):
        """expand_panel 안에 종목별 행 다시 그림 (기존 layout 폐기 후 재구성)."""
        # 기존 layout 정리 (dummy QWidget로 양도 → GC)
        old = self.expand_panel.layout()
        if old is not None:
            QWidget().setLayout(old)

        vl = QVBoxLayout(self.expand_panel)
        vl.setContentsMargins(14, self.PANEL_TOP, 14, self.PANEL_BOTTOM)
        vl.setSpacing(self.ROW_SPACING)

        # 상단 구분선
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background: {C['border']}; max-height: 1px; border: none;")
        vl.addWidget(sep)

        for h in self.holdings:
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(8)

            name_lbl = QLabel(h["name"])
            name_lbl.setStyleSheet(f"color: {C['text']}; font-size: 11px;")
            row.addWidget(name_lbl, 1)

            profit = int(h["profit"])
            rate   = float(h["profit_rate"])
            if profit > 0:
                color, sign = C['red'], "+"
            elif profit < 0:
                color, sign = C['blue'], ""   # 음수는 자체 '-' 사용
            else:
                color, sign = C['subtext'], ""

            profit_lbl = QLabel(f"{sign}{profit:,} 원")
            profit_lbl.setStyleSheet(f"color: {color}; font-size: 11px; font-weight: bold;")
            profit_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            profit_lbl.setFixedWidth(100)
            row.addWidget(profit_lbl)

            rate_lbl = QLabel(f"{sign}{rate:.2f}%")
            rate_lbl.setStyleSheet(f"color: {color}; font-size: 11px;")
            rate_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            rate_lbl.setFixedWidth(60)
            row.addWidget(rate_lbl)

            vl.addLayout(row)

    # ── 확장 / 축소 토글 ─────────────────────────────────────────────────
    def toggle_expand(self):
        if self.is_expanded:
            self.collapse()
        else:
            self.expand()

    def expand(self):
        if self.is_expanded or not self.holdings:
            return
        self.is_expanded = True
        self._render_holdings()
        self._resize_to_expanded()
        self.expand_panel.show()

    def collapse(self):
        if not self.is_expanded:
            return
        self.is_expanded = False
        self.expand_panel.hide()
        self.setFixedHeight(self.H)
        self.card.setGeometry(0, 0, self.W, self.H)

    # ── 드래그 이동 + 클릭 토글 (StockWidget 와 동일 패턴) ────────────────
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos  = event.globalPosition().toPoint() - self.pos()
            self._press_pos = event.globalPosition().toPoint()
            self._moved     = False

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.MouseButton.LeftButton and self._drag_pos:
            if not self._moved and self._press_pos:
                delta = event.globalPosition().toPoint() - self._press_pos
                if abs(delta.x()) > self.DRAG_THRESHOLD or abs(delta.y()) > self.DRAG_THRESHOLD:
                    self._moved = True
            if self._moved:
                self.move(event.globalPosition().toPoint() - self._drag_pos)

    def mouseReleaseEvent(self, event):
        # 드래그가 아니었으면 = 클릭 → 종목 목록 토글
        if event.button() == Qt.MouseButton.LeftButton and not self._moved:
            self.toggle_expand()
        self._drag_pos  = None
        self._press_pos = None
        self._moved     = False


# ─── 전체 위젯 관리자 ─────────────────────────────────────────────────────────
class WidgetManager:
    def __init__(self, app: QApplication):
        self.app = app
        self.stocks: list[dict] = []
        self.widgets: dict[str, StockWidget] = {}
        self.uniform_w: int = StockWidget.MIN_W
        self.is_hidden: bool = False    # 위젯 전체 숨김 상태
        # 마스터 위젯 (포트폴리오 요약)
        self.master_widget: MasterWidget | None = None
        self.master_visible: bool = True
        self.master_pos: list | None = None   # None → 기본 위치

        self._load_config()
        self._setup_tray()
        self._spawn_all()

    # ── 전체 위젯 표시/숨김 토글 ─────────────────────────────────────────
    def toggle_visibility(self):
        self.is_hidden = not self.is_hidden
        for w in self.widgets.values():
            w.hide() if self.is_hidden else w.show()
        # 마스터 위젯도 전체 토글에 함께 따름. 단, 마스터 개별 숨김 상태는 보존.
        if self.master_widget:
            if self.is_hidden:
                self.master_widget.hide()
            elif self.master_visible:
                self.master_widget.show()
        self.toggle_act.setText("👀   표시하기" if self.is_hidden else "🙈   숨기기")

    # ── 위치 초기화 ───────────────────────────────────────────────────────
    def reset_positions(self):
        """각 위젯을 현재 위치한 모니터의 우상단부터 column-wrap 방식으로 정렬.
        - 첫 column이 화면 세로 영역을 넘어가면 그 왼쪽에 새 column을 시작
        - 마스터 위젯이 표시 중이면 자기 모니터의 우상단 첫 자리에 두고,
          모든 column은 마스터 아래 y부터 시작 (마스터보다 위로는 가지 않음)"""
        MARGIN_X      = 20   # 화면 우측 여백
        MARGIN_Y      = 60   # 화면 상단 여백
        MARGIN_BOTTOM = 20   # 화면 하단 여백 (이 안쪽으로만 위젯 배치)
        GAP           = 4    # 같은 column 내 위젯 간 세로 간격
        COL_GAP       = 8    # column 사이 가로 간격

        # 마스터 위젯이 표시 중인 모니터 파악
        master_screen = None
        master_offset = 0
        if self.master_widget and self.master_widget.isVisible():
            mc = self.master_widget.frameGeometry().center()
            master_screen = QApplication.screenAt(mc) or QApplication.primaryScreen()
            mgeo = master_screen.availableGeometry()
            mx = mgeo.x() + mgeo.width() - self.master_widget.width() - MARGIN_X
            my = mgeo.y() + MARGIN_Y
            self.master_widget.move(mx, my)
            self.master_pos = [mx, my]
            master_offset = self.master_widget.height() + GAP

        # 위젯을 현재 속한 모니터별로 그룹화 (stocks 순서 보존)
        groups: dict = {}
        for s in self.stocks:
            w = self.widgets.get(s["code"])
            if not w:
                continue
            center = w.frameGeometry().center()
            screen = QApplication.screenAt(center) or QApplication.primaryScreen()
            groups.setdefault(screen, []).append((s, w))

        widget_w = self.uniform_w
        step_y   = StockWidget.COMPACT_H + GAP

        for screen, items in groups.items():
            geo = screen.availableGeometry()
            col_top_y = geo.y() + MARGIN_Y + (master_offset if screen is master_screen else 0)
            # 한 column에 들어가는 위젯 수 (하단 여백까지 고려)
            avail_h = geo.y() + geo.height() - MARGIN_BOTTOM - col_top_y
            max_per_col = max(1, avail_h // step_y)

            first_col_x = geo.x() + geo.width() - widget_w - MARGIN_X
            for i, (s, w) in enumerate(items):
                col_idx = i // max_per_col
                row_idx = i %  max_per_col
                x = first_col_x - col_idx * (widget_w + COL_GAP)
                y = col_top_y + row_idx * step_y
                w.move(x, y)
                s["pos"] = [x, y]

        self._save_config()
        # 숨김 상태라면 자동으로 다시 표시
        if self.is_hidden:
            self.toggle_visibility()

    # ── 통일 너비 계산/적용 ───────────────────────────────────────────────
    def _calc_uniform_width(self) -> int:
        """모든 종목명 중 가장 긴 이름 기준 통일 너비."""
        w = StockWidget.MIN_W
        for s in self.stocks:
            name = s.get("name", s["code"])
            w = max(w, StockWidget.calc_width_for_name(name))
        return w

    def _apply_uniform_width(self):
        """현재 너비를 재계산해 모든 위젯에 적용."""
        new_w = self._calc_uniform_width()
        if new_w == self.uniform_w:
            return
        self.uniform_w = new_w
        for w in self.widgets.values():
            w.set_width(new_w)
        if self.master_widget:
            self.master_widget.set_uniform_width(new_w)

    # ── 트레이 ─────────────────────────────────────────────────────────────
    def _setup_tray(self):
        icon = self._make_tray_icon()
        self.tray = QSystemTrayIcon(icon)
        self.tray.setToolTip("한국 주식 위젯")

        menu = QMenu()
        menu.setStyleSheet(TRAY_MENU_STYLE)

        add_act    = QAction("➕   종목 추가",   menu)
        manage_act = QAction("📋   종목 관리",   menu)
        export_act = QAction("📤   Excel로 내보내기", menu)
        import_act = QAction("📥   Excel에서 가져오기", menu)
        self.toggle_act = QAction("🙈   숨기기", menu)
        self.master_toggle_act = QAction(self._master_toggle_text(), menu)
        reset_act  = QAction("📐   위치 초기화", menu)
        quit_act   = QAction("❌   종료",        menu)
        add_act.triggered.connect(self.open_add_dialog)
        manage_act.triggered.connect(self.open_manage_dialog)
        export_act.triggered.connect(self.open_export_dialog)
        import_act.triggered.connect(self.open_import_dialog)
        self.toggle_act.triggered.connect(self.toggle_visibility)
        self.master_toggle_act.triggered.connect(self.toggle_master_visibility)
        reset_act.triggered.connect(self.reset_positions)
        quit_act.triggered.connect(self.app.quit)

        menu.addAction(add_act)
        menu.addAction(manage_act)
        menu.addSeparator()
        menu.addAction(export_act)
        menu.addAction(import_act)
        menu.addSeparator()
        menu.addAction(self.toggle_act)
        menu.addAction(self.master_toggle_act)
        menu.addAction(reset_act)
        menu.addSeparator()
        menu.addAction(quit_act)

        self.tray.setContextMenu(menu)
        self.tray.activated.connect(self._on_tray_activated)
        self.tray.show()

    def _on_tray_activated(self, reason):
        # 트레이 아이콘 좌클릭(Trigger) 시 표시/숨김 빠른 토글
        if reason == QSystemTrayIcon.ActivationReason.Trigger:
            self.toggle_visibility()

    @staticmethod
    def _make_tray_icon() -> QIcon:
        px = QPixmap(32, 32)
        px.fill(QColor(0, 0, 0, 0))
        p = QPainter(px)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setBrush(QBrush(QColor(C["blue"])))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(1, 1, 30, 30)
        p.setFont(QFont("Malgun Gothic",14, QFont.Weight.Bold))
        p.setPen(QPen(QColor(C["bg"])))
        p.drawText(px.rect(), Qt.AlignmentFlag.AlignCenter, "₩")
        p.end()
        return QIcon(px)

    # ── 설정 파일 ──────────────────────────────────────────────────────────
    # 스키마 변천:
    #   v1 (구버전): JSON 루트가 list — 종목 dict 의 배열
    #   v2 (현재):   JSON 루트가 dict — {"stocks": [...], "master": {"visible": bool, "pos": [x,y]|null}}
    # 로드는 둘 다 받아주고, 저장은 항상 v2 로 한다 (한 번 저장되면 자동 마이그레이트).
    def _load_config(self):
        if not os.path.exists(CONFIG_FILE):
            return
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return

        if isinstance(data, list):
            # v1 → 종목만 있음, 마스터 설정은 기본값
            self.stocks = data
        elif isinstance(data, dict):
            self.stocks = data.get("stocks", []) or []
            master = data.get("master") or {}
            self.master_visible = bool(master.get("visible", True))
            pos = master.get("pos")
            if isinstance(pos, list) and len(pos) == 2:
                try:
                    self.master_pos = [int(pos[0]), int(pos[1])]
                except (TypeError, ValueError):
                    self.master_pos = None

    def _save_config(self):
        data = {
            "stocks": self.stocks,
            "master": {
                "visible": self.master_visible,
                "pos": self.master_pos,
            },
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[save] 오류: {e}")

    def save_positions(self):
        for s in self.stocks:
            w = self.widgets.get(s["code"])
            if w:
                pos = w.pos()
                s["pos"] = [pos.x(), pos.y()]
        if self.master_widget:
            mpos = self.master_widget.pos()
            self.master_pos = [mpos.x(), mpos.y()]
        self._save_config()

    # ── 위젯 생성 ──────────────────────────────────────────────────────────
    def _spawn_all(self):
        self.uniform_w = self._calc_uniform_width()
        for i, s in enumerate(self.stocks):
            default_x = 60
            default_y = 60 + i * (StockWidget.COMPACT_H + 12)
            self._spawn_widget(s, default_x, default_y, stagger_idx=i)
        self._spawn_master()

    def _spawn_widget(self, stock: dict, def_x=60, def_y=60, stagger_idx: int = 0):
        code = stock["code"]
        w = StockWidget(stock, width=self.uniform_w, stagger_idx=stagger_idx)
        w.deleted.connect(self._on_delete)
        w.edited.connect(self._on_edited)
        w.price_updated.connect(lambda _: self._recompute_master())

        pos = stock.get("pos", [def_x, def_y])
        w.move(pos[0], pos[1])
        w.show()
        self.widgets[code] = w

    def _on_edited(self, _code: str):
        """개별 위젯에서 평단가/수량을 수정한 경우. 저장 + 마스터 갱신."""
        self._save_config()
        self._recompute_master()

    # ── 마스터 위젯 생성/표시 ─────────────────────────────────────────────
    def _spawn_master(self):
        if self.master_widget is None:
            self.master_widget = MasterWidget(width=self.uniform_w)

        # 위치: 저장된 위치가 있으면 사용, 없으면 종목 위젯들 위에 적당히 둠
        if self.master_pos:
            self.master_widget.move(self.master_pos[0], self.master_pos[1])
        else:
            self.master_widget.move(60, 20)

        if self.master_visible and not self.is_hidden:
            self.master_widget.show()
        else:
            self.master_widget.hide()

        # 초기 표시: 현재가 아직 없으면 0/─ 으로 둠 → 30초 이내 자동 갱신
        self._recompute_master()

    def _master_toggle_text(self) -> str:
        return "📊   마스터 위젯 숨기기" if self.master_visible else "📊   마스터 위젯 표시"

    def toggle_master_visibility(self):
        self.master_visible = not self.master_visible
        if self.master_widget:
            if self.master_visible and not self.is_hidden:
                self.master_widget.show()
            else:
                self.master_widget.hide()
        self.master_toggle_act.setText(self._master_toggle_text())
        self._save_config()

    def _recompute_master(self):
        """모든 종목 위젯의 current_price 를 모아 마스터 4지표 및 보유 종목 상세를 갱신."""
        if not self.master_widget:
            return
        if not self.stocks:
            self.master_widget.clear_metrics()
            return

        total_invest = 0
        total_eval   = 0
        holdings: list[dict] = []
        for s in self.stocks:
            avg = int(s.get("avg_price", 0))
            qty = int(s.get("quantity", 0))
            invest = avg * qty
            total_invest += invest

            w = self.widgets.get(s["code"])
            # 현재가가 아직 안 잡힌 종목은 평가금액에서 평단가로 임시 사용
            price = w.current_price if (w and w.current_price) else avg
            eval_v = price * qty
            total_eval += eval_v

            profit = eval_v - invest
            rate   = (profit / invest * 100.0) if invest else 0.0
            holdings.append({
                "name":        s.get("name", s["code"]),
                "profit":      profit,
                "profit_rate": rate,
            })

        self.master_widget.update_metrics(total_invest, total_eval)
        self.master_widget.update_holdings(holdings)

    # ── 종목 추가 ──────────────────────────────────────────────────────────
    def open_add_dialog(self):
        dlg = StockDialog()
        if not dlg.exec():
            return
        d = dlg.get_data()
        code = d["code"]

        if not code:
            return
        if code in self.widgets:
            QMessageBox.information(None, "알림", f"'{code}'는 이미 추가되어 있습니다.")
            return

        # 종목명 미리 조회
        result = fetch_stock(code)
        if not result:
            QMessageBox.warning(None, "조회 실패", f"종목코드 '{code}'를 찾을 수 없습니다.\n코드를 다시 확인해 주세요.")
            return

        d["name"] = result["name"]
        self.stocks.append(d)
        self._save_config()

        # 새 종목명이 더 길면 모든 위젯 너비 재조정 (새 위젯도 이 값으로 생성됨)
        self._apply_uniform_width()

        # 새 위젯 위치: 기존 위젯들 아래. 추가된 위젯이라 stagger 필요 없음(즉시 시작)
        ny = 60 + len(self.widgets) * (StockWidget.COMPACT_H + 12)
        self._spawn_widget(d, 60, ny, stagger_idx=0)

        self._recompute_master()

        # 숨김 상태에서 새 종목을 추가한 경우 자동으로 표시 상태로 전환
        if self.is_hidden:
            self.toggle_visibility()

    # ── 종목 일괄 관리 ────────────────────────────────────────────────────
    def open_manage_dialog(self):
        # 평가손익 계산용 현재가 스냅샷
        current_prices = {
            code: int(w.current_price)
            for code, w in self.widgets.items()
            if w.current_price
        }
        dlg = ManageStocksDialog(
            stocks=copy.deepcopy(self.stocks),
            current_prices=current_prices,
        )
        if not dlg.exec():
            return
        new_stocks = dlg.get_stocks()

        old_map = {s["code"]: s for s in self.stocks}
        new_map = {s["code"]: s for s in new_stocks}

        # 삭제된 종목: 위젯 닫고 제거
        for code in list(old_map):
            if code not in new_map:
                w = self.widgets.pop(code, None)
                if w:
                    w.close()

        # 추가된 종목: 위젯 생성 (기본 위치) — 다수 추가 시 stagger로 분산
        added_idx = 0
        for s in new_stocks:
            if s["code"] not in old_map:
                ny = 60 + len(self.widgets) * (StockWidget.COMPACT_H + 12)
                self._spawn_widget(s, 60, ny, stagger_idx=added_idx)
                added_idx += 1

        # 기존 종목: 평단가/수량 변경 반영
        for s in new_stocks:
            code = s["code"]
            if code in old_map and code in self.widgets:
                w = self.widgets[code]
                w.data["avg_price"] = s["avg_price"]
                w.data["quantity"]  = s["quantity"]
                if w.current_price:
                    w._update_detail(w.current_price)

        # 순서 + 저장 + 너비 재계산
        self.stocks = new_stocks
        self._apply_uniform_width()
        self._save_config()
        self._recompute_master()

        # 숨김 상태에서 변경된 종목이 있으면 자동으로 표시 상태로 전환
        if self.is_hidden and self.widgets:
            self.toggle_visibility()

    # ── Excel 내보내기 ────────────────────────────────────────────────────
    def open_export_dialog(self):
        if not self.stocks:
            QMessageBox.information(None, "알림", "내보낼 보유 종목이 없습니다.")
            return

        default_name = f"pinstock_holdings_{datetime.now().strftime('%Y%m%d')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            None, "보유 종목 Excel로 내보내기",
            os.path.join(os.path.expanduser("~"), default_name),
            "Excel 파일 (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # 마스터 위젯과 동일한 4지표를 시트 하단에 포함시키기 위해 현재가 dict 전달
        current_prices = {
            code: w.current_price
            for code, w in self.widgets.items()
            if w.current_price
        }

        try:
            export_stocks_to_excel(self.stocks, path, current_prices)
        except ImportError:
            QMessageBox.critical(
                None, "라이브러리 없음",
                "openpyxl 패키지가 필요합니다.\n\n터미널에서 다음을 실행하세요:\n    pip install openpyxl"
            )
            return
        except Exception as e:
            QMessageBox.critical(None, "내보내기 실패", f"파일을 저장할 수 없습니다.\n\n{e}")
            return

        QMessageBox.information(
            None, "내보내기 완료",
            f"{len(self.stocks)}개 종목을 저장했습니다.\n\n{path}"
        )

    # ── Excel 가져오기 ────────────────────────────────────────────────────
    def open_import_dialog(self):
        path, _ = QFileDialog.getOpenFileName(
            None, "Excel에서 보유 종목 가져오기",
            os.path.expanduser("~"),
            "Excel 파일 (*.xlsx)"
        )
        if not path:
            return

        try:
            imported = import_stocks_from_excel(path)
        except ImportError:
            QMessageBox.critical(
                None, "라이브러리 없음",
                "openpyxl 패키지가 필요합니다.\n\n터미널에서 다음을 실행하세요:\n    pip install openpyxl"
            )
            return
        except ValueError as e:
            QMessageBox.critical(None, "가져오기 실패", str(e))
            return
        except Exception as e:
            QMessageBox.critical(None, "가져오기 실패", f"파일을 읽을 수 없습니다.\n\n{e}")
            return

        # 모드 선택
        mode_dlg = ImportModeDialog()
        if not mode_dlg.exec():
            return
        mode = mode_dlg.mode

        # 미리보기 / 최종 확인
        if mode == "overwrite":
            msg = (
                f"덮어쓰기 모드입니다.\n\n"
                f"기존 {len(self.stocks)}개 종목이 모두 삭제되고\n"
                f"Excel의 {len(imported)}개 종목으로 교체됩니다.\n\n"
                "계속할까요?"
            )
        else:
            new_codes = {s["code"] for s in imported}
            existing_codes = {s["code"] for s in self.stocks}
            updated = len(new_codes & existing_codes)
            added = len(new_codes - existing_codes)
            msg = (
                f"병합 모드입니다.\n\n"
                f"• 갱신: {updated}개 (기존 종목 평단가/수량 업데이트)\n"
                f"• 추가: {added}개 (새 종목)\n"
                f"• 유지: {len(existing_codes - new_codes)}개 (Excel에 없는 기존 종목)\n\n"
                "계속할까요?"
            )
        ret = QMessageBox.question(
            None, "가져오기 확인", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if ret != QMessageBox.StandardButton.Yes:
            return

        # 적용 직전에 위치 저장 (병합 모드에서 기존 위치 보존하려면 최신 좌표가 필요)
        self.save_positions()

        # stocks.json 백업
        if os.path.exists(CONFIG_FILE):
            try:
                shutil.copy2(CONFIG_FILE, BACKUP_FILE)
            except Exception as e:
                print(f"[backup] 오류: {e}")

        # 새 stocks 리스트 구성
        if mode == "overwrite":
            new_stocks = imported   # pos 없음 → 다시 spawn 시 기본 위치
        else:
            by_code = {s["code"]: s for s in self.stocks}
            new_stocks = []
            for s in imported:
                # 기존 항목이 있으면 pos 등 부가 정보 보존
                base = dict(by_code.get(s["code"], {}))
                base.update(s)   # 평단가/수량/이름은 Excel 값으로 갱신
                new_stocks.append(base)
            # Excel 에 없는 기존 종목은 뒤에 그대로 유지
            imported_codes = {s["code"] for s in imported}
            for s in self.stocks:
                if s["code"] not in imported_codes:
                    new_stocks.append(s)

        self._rebuild_widgets(new_stocks)

        QMessageBox.information(
            None, "가져오기 완료",
            f"총 {len(new_stocks)}개 종목이 적용되었습니다.\n"
            f"이전 데이터는 다음에 백업되었습니다:\n{BACKUP_FILE}"
        )

    # ── 종목 리스트 전체 교체 후 위젯 재구성 ─────────────────────────────
    def _rebuild_widgets(self, new_stocks: list[dict]):
        """기존 위젯을 모두 닫고 new_stocks 기준으로 위젯을 다시 생성한다."""
        for w in list(self.widgets.values()):
            w.close()
        self.widgets.clear()

        self.stocks = new_stocks
        self.uniform_w = self._calc_uniform_width()

        for i, s in enumerate(self.stocks):
            default_x = 60
            default_y = 60 + i * (StockWidget.COMPACT_H + 12)
            self._spawn_widget(s, default_x, default_y, stagger_idx=i)

        # 마스터 위젯도 새 너비에 맞춰 갱신
        if self.master_widget:
            self.master_widget.set_uniform_width(self.uniform_w)

        self._save_config()
        self._recompute_master()

        # 위치 정보가 없는 종목들이 있으면 자동으로 정렬
        if any("pos" not in s for s in self.stocks):
            self.reset_positions()

        if self.is_hidden and self.widgets:
            self.toggle_visibility()

    # ── 종목 삭제 ──────────────────────────────────────────────────────────
    def _on_delete(self, code: str):
        self.stocks = [s for s in self.stocks if s["code"] != code]
        self.widgets.pop(code, None)
        self._save_config()
        # 가장 긴 종목이 삭제된 경우 남은 위젯들도 줄어들도록
        self._apply_uniform_width()
        self._recompute_master()


# ─── 진입점 ───────────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)  # 트레이만 있어도 계속 실행

    manager = WidgetManager(app)
    app.aboutToQuit.connect(manager.save_positions)

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
